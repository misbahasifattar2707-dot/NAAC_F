import os
import io
import re
import uuid
from flask import Blueprint, jsonify, request, session, current_app, send_file
from werkzeug.utils import secure_filename
from models.models import *
from extensions import db, bcrypt
import pandas as pd
from datetime import datetime, date
from decimal import Decimal
from urllib.parse import urlparse, unquote

# Excel + PDF + Image utilities
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfWriter, PdfReader
from PIL import Image

from routes.criteria346_merge import (
    CRITERIA_346_MODELS,
    CRITERION_346_TITLES,
    enrich_criteria346_to_dict,
    get_records_c346,
    apply_c346_foreign_keys,
    apply_c346_column_mapping,
    apply_c346_update,
    register_c346_routes,
    DB_TO_UI,
)

api_bp = Blueprint('api_bp', __name__, url_prefix='/api')

# --- Auth Endpoints ---

@api_bp.route('/register', methods=['POST'])
def api_register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'Admin')
    department = data.get('department', '')
    user_identifier = data.get('user_identifier', '')
    year = data.get('year', '')

    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({"success": False, "error": "Username already exists"}), 400

    hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
    new_user = User(
        username=username, 
        password=hashed_pw, 
        role=role, 
        department=department,
        user_identifier=user_identifier,
        year=year
    )
    db.session.add(new_user)
    db.session.commit()
    
    return jsonify({"success": True, "message": "Account created"})


@api_bp.route('/login', methods=['POST'])
def api_login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    academic_year = data.get('academic_year')
    
    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400
        
    user = User.query.filter_by(username=username).first()
    if user and bcrypt.check_password_hash(user.password, password):
        session['user_id'] = user.id
        session['academic_year'] = academic_year
        
        return jsonify({
            "success": True, 
            "user": {
                "id": user.id, 
                "username": user.username, 
                "role": user.role, 
                "department": user.department,
                "academic_year": academic_year or user.year,
                "program": "MCA",
                "programCode": "515124110"
            }
        })
    return jsonify({"success": False, "error": "Invalid username or password"}), 401


# --- Metadata Endpoints ---

@api_bp.route('/teachers', methods=['GET'])
def get_teachers():
    """All teachers from teacher_lookup for Criteria 1.1.3, 2.4.1, 2.4.2 dropdowns."""
    teachers = Teacher.query.order_by(Teacher.name).all()
    return jsonify([
        {
            "id": t.id,
            "name": t.name,
            "pan": t.pan or "",
            "designation": t.designation or "",
        }
        for t in teachers
    ])


@api_bp.route('/teachers', methods=['POST'])
def add_teacher():
    """Add a teacher row for dropdowns (Criterion 2.4.x). Minimum: name."""
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": "Name is required"}), 400
    pan = (data.get("pan") or "").strip() or None
    designation = (data.get("designation") or "").strip() or None
    if pan and Teacher.query.filter_by(pan=pan).first():
        return jsonify({"success": False, "error": "A teacher with this PAN already exists"}), 400
    try:
        t = Teacher(name=name, pan=pan, designation=designation)
        db.session.add(t)
        db.session.commit()
        return jsonify({
            "success": True,
            "teacher": {"id": t.id, "name": t.name, "pan": t.pan or "", "designation": t.designation or ""},
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/departments', methods=['GET'])
def get_departments():
    progs = Program.query.all()
    res, seen_codes = [], set()
    for p in progs:
        pc = (p.program_code or '').strip()
        if not pc or pc in seen_codes:
            continue
        seen_codes.add(pc)
        disp = _program_display_name(p)
        label = (p.department or '').strip() or disp
        res.append({"id": p.id, "code": label, "programCode": p.program_code, "programName": disp})
    if not res:
        res = [{"id": 1, "code": "MCA", "programCode": "515124110", "programName": "MCA"}]
    return jsonify(res)

@api_bp.route('/semesters', methods=['GET'])
def get_semesters():
    recs = SemesterLookup.query.all()
    if not recs:
        # Default seed
        defaults = ["FYMCA-SEM-I (MCA)", "FYMCA-SEM-II (MCA)", "SYMCA-SEM-III (MCA)", "SYMCA-SEM-IV (MCA)"]
        for d in defaults: db.session.add(SemesterLookup(value=d))
        db.session.commit()
        recs = SemesterLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/courses', methods=['GET'])
def get_courses():
    """
    GET /api/courses          → all courses in SemesterCourseLookup
    GET /api/courses?sem=X    → courses for that semester only
    """
    sem = request.args.get('sem')
    q = SemesterCourseLookup.query
    if sem:
        q = q.filter(SemesterCourseLookup.semester == sem)
    rows = q.order_by(SemesterCourseLookup.course_code).all()
    return jsonify([{"code": r.course_code, "name": r.course_name, "semester": r.semester} for r in rows])

@api_bp.route('/courses', methods=['POST'])
def add_course():
    """
    POST /api/courses  { semester, course_code, course_name }
    Adds a new course to SemesterCourseLookup with normalisation + uniqueness check.
    """
    data = request.json or {}
    sem    = (data.get('semester') or '').strip()
    c_code = (data.get('course_code') or '').strip().upper()
    c_name = (data.get('course_name') or '').strip()

    if not sem or not c_code or not c_name:
        return jsonify({"success": False, "error": "semester, course_code and course_name are required"}), 400

    existing = SemesterCourseLookup.query.filter_by(semester=sem, course_code=c_code).first()
    if existing:
        return jsonify({"success": False, "error": f"Course {c_code} already exists in {sem}"}), 400

    try:
        new_c = SemesterCourseLookup(semester=sem, course_code=c_code, course_name=c_name)
        db.session.add(new_c)
        db.session.commit()
        return jsonify({"success": True, "course": {"code": c_code, "name": c_name, "semester": sem}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400


@api_bp.route('/electives', methods=['GET'])
def get_electives():
    # Attempting to load by 310916% or elective
    courses = Course.query.filter((Course.course_code.like('310916%')) | (Course.course_name.ilike('%elective%'))).all()
    return jsonify([{"id": c.id, "code": c.course_code, "name": c.course_name} for c in courses])


@api_bp.route('/electives', methods=['POST'])
def add_elective_course():
    """Add a Course row shown under /api/electives (Criterion 1.2.1 subject codes)."""
    data = request.json or {}
    code = (data.get("code") or data.get("course_code") or "").strip().upper()
    name = (data.get("name") or data.get("course_name") or "").strip()
    if not code or not name:
        return jsonify({"success": False, "error": "code and name are required"}), 400
    if Course.query.filter_by(course_code=code).first():
        return jsonify({"success": False, "error": "This course code already exists"}), 400
    try:
        c = Course(course_code=code, course_name=name)
        db.session.add(c)
        db.session.commit()
        return jsonify({"success": True, "elective": {"id": c.id, "code": c.course_code, "name": c.course_name}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400


@api_bp.route('/addon-durations', methods=['GET'])
def get_addon_durations():
    """Criterion 1.2.2 duration labels; seeded like semesters when empty."""
    recs = AddonDurationLookup.query.all()
    if not recs:
        defaults = [f"{i} Month{'s' if i > 1 else ''}" for i in range(1, 13)]
        for d in defaults:
            db.session.add(AddonDurationLookup(value=d))
        db.session.commit()
        recs = AddonDurationLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/program-codes', methods=['GET'])
def get_program_codes():
    progs = Program.query.all()
    out = []
    for p in progs:
        disp = _program_display_name(p)
        pname = (p.program_name or "").strip()
        out.append({
            "code": p.program_code,
            "label": disp or pname or p.program_code,
            "programName": pname or disp or p.program_code,
            "department": (p.department or "").strip(),
        })
    return jsonify(out)

@api_bp.route('/course-types', methods=['GET'])
def get_course_types():
    recs = CourseTypeLookup.query.all()
    if not recs:
        defaults = ["PBL (Project Based Learning)", "Major Project"]
        for d in defaults:
            db.session.add(CourseTypeLookup(value=d))
        db.session.commit()
        recs = CourseTypeLookup.query.all()
    out = []
    changed = False
    for x in recs:
        cc = (getattr(x, "course_code", None) or "").strip() or _guess_course_code_from_type_label(x.value)
        if not getattr(x, "course_code", None) and cc:
            x.course_code = cc
            changed = True
        out.append({"value": x.value, "label": x.value, "courseCode": cc or ""})
    if changed:
        db.session.commit()
    return jsonify(out)

@api_bp.route('/academic-years', methods=['GET'])
def get_academic_years():
    recs = AcademicYearLookup.query.all()
    if not recs:
        defaults = ["2020-21", "2021-22", "2022-23", "2023-24", "2024-25", "2025-26", "2026-27"]
        for d in defaults: db.session.add(AcademicYearLookup(value=d))
        db.session.commit()
        recs = AcademicYearLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/programmes', methods=['GET'])
def get_programmes():
    progs = Program.query.all()
    if not progs:
        return jsonify([
            {"code": "MCA", "name": "Master of Computer Applications (MCA)", "department": "", "display_name": "MCA"},
            {"code": "MBA", "name": "Master of Business Administration (MBA)", "department": "", "display_name": "MBA"},
        ])
    return jsonify([
        {
            "code": p.program_code,
            "name": p.program_name,
            "department": p.department or "",
            "display_name": _program_display_name(p),
        }
        for p in progs
    ])


@api_bp.route('/programmes', methods=['POST'])
def add_programme():
    """Create a programme for lookups / Criteria 2.1.1 dropdown."""
    data = request.json or {}
    code = (data.get('program_code') or data.get('code') or '').strip()
    name = (data.get('program_name') or data.get('name') or '').strip()
    if not code or not name:
        return jsonify({"success": False, "error": "program_code and program_name are required"}), 400
    if Program.query.filter_by(program_code=code).first():
        return jsonify({"success": False, "error": "This programme code already exists"}), 400
    user_id = session.get('user_id')
    dept_label = (data.get("department") or "").strip() or None

    prog = Program(
        program_code=code,
        program_name=name,
        department=dept_label,
        created_by_id=user_id,
        updated_by_id=user_id,
    )
    db.session.add(prog)
    db.session.commit()
    return jsonify({
        "success": True,
        "program": {
            "code": prog.program_code,
            "name": prog.program_name,
            "department": prog.department or "",
            "display_name": _program_display_name(prog),
        },
    })


@api_bp.route('/reserved-categories', methods=['GET'])
def get_reserved_categories():
    recs = ReservedCategoryLookup.query.all()
    if not recs:
        defaults = ["SC", "ST", "OBC", "Divyangjan", "Gen-EWS", "Others"]
        for d in defaults: db.session.add(ReservedCategoryLookup(value=d))
        db.session.commit()
        recs = ReservedCategoryLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/library-resources', methods=['GET'])
def get_library_resources():
    recs = LibraryResourceLookup.query.all()
    if not recs:
        defaults = ["Books", "Journals", "e-Journals", "e-Books", "e-ShodhSindhu", "Shodhganga", "Databases"]
        for d in defaults: db.session.add(LibraryResourceLookup(value=d))
        db.session.commit()
        recs = LibraryResourceLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/qualifying-exams', methods=['GET'])
def get_qualifying_exams():
    recs = QualifyingExamLookup.query.all()
    if not recs:
        defaults = ["NET", "SLET", "GATE", "GMAT", "CAT", "GRE", "JAM", "IELTS", "TOEFL", "Civil Services", "State government examinations", "Other examinations"]
        for d in defaults: db.session.add(QualifyingExamLookup(value=d))
        db.session.commit()
        recs = QualifyingExamLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/event-levels', methods=['GET'])
def get_event_levels():
    recs = EventLevelLookup.query.all()
    if not recs:
        defaults = ["Inter-university", "State", "National", "International", "District"]
        for d in defaults: db.session.add(EventLevelLookup(value=d))
        db.session.commit()
        recs = EventLevelLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/award-categories', methods=['GET'])
def get_award_categories():
    recs = AwardCategoryLookup.query.all()
    if not recs:
        defaults = ["Individual", "Team"]
        for d in defaults: db.session.add(AwardCategoryLookup(value=d))
        db.session.commit()
        recs = AwardCategoryLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

# --- Generic POST Endpoint to add new lookup items ---
LOOKUP_MODELS = {
    "semesters": SemesterLookup,
    "course-types": CourseTypeLookup,
    "academic-years": AcademicYearLookup,
    "reserved-categories": ReservedCategoryLookup,
    "library-resources": LibraryResourceLookup,
    "qualifying-exams": QualifyingExamLookup,
    "event-levels": EventLevelLookup,
    "award-categories": AwardCategoryLookup,
    "departments": DepartmentLookup,
    "designations": DesignationLookup,
    "highest-degrees": HighestDegreeLookup,
    "appointment-types": AppointmentTypeLookup,
    "funding-agencies": FundingAgencyLookup,
    "levels": LevelLookup,
    "team-individual": TeamIndividualLookup,
    "addon-durations": AddonDurationLookup,
    "egovernance-areas": EGovernanceAreaLookup
}

# --- Dedicated Getters for new Lookups ---

@api_bp.route('/get-lookups/<lookup_key>', methods=['GET'])
def get_lookup_values(lookup_key):
    model = LOOKUP_MODELS.get(lookup_key)
    if not model: return jsonify([])
    recs = model.query.all()
    # Handle different models if needed, but most use .value
    return jsonify([{"id": r.id, "value": r.value, "label": r.value} for r in recs])

@api_bp.route('/lookups/<lookup_key>', methods=['POST'])
def add_lookup_value(lookup_key):
    model = LOOKUP_MODELS.get(lookup_key)
    if not model:
        return jsonify({"success": False, "error": "Invalid lookup key"}), 404
        
    data = request.json or {}
    new_val = data.get('value')
    if not new_val:
        return jsonify({"success": False, "error": "Value required"}), 400
        
    existing = model.query.filter_by(value=new_val).first()
    if not existing:
        try:
            kwargs = {"value": new_val}
            if lookup_key == "course-types" and hasattr(model, "course_code"):
                cc = (data.get("course_code") or "").strip() or _guess_course_code_from_type_label(new_val)
                if cc:
                    kwargs["course_code"] = cc
            new_item = model(**kwargs)
            db.session.add(new_item)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 400
            
    return jsonify({"success": True, "message": f"Added {new_val}"})

# --- Generic Criteria Router ---

CRITERIA_MODELS = {
    "1_1": C11Courses, "1_1_3": C113TeacherBodies, "1_2_1": C121CBCS, "1_2_2": C122Addon,
    "1_3_2": C132Experiential, "1_3_3": C133Projects,
    "2_1": C21StudentsDuringYear, "2_2": C22ReservedSeats,
    "2_1_1": C211Enrolment, "2_1_2": C212Reservation, "2_3": C23OutgoingStudents,
    "2_3_3": C233MentorRatio, "2_3_3_meta": C233MentorMeta,
    "2_4_1": C241Teachers, "2_4_2": C242TeacherPhD, "2_6_3": C263PassPercentage,
    **CRITERIA_346_MODELS,
}


def _program_display_name(program):
    """Degree/programme label for NAAC UI (e.g. MCA), not semester rows like FYMCA SEM I."""
    if not program:
        return ''
    code = (program.program_code or '').strip().upper()
    pname = (program.program_name or '').strip()
    pname_norm = pname.upper().replace(' ', '').replace('-', '')
    if code in ('515124110', 'MCA'):
        return 'MCA'
    if 'FYMCA' in pname_norm or 'SYMCA' in pname_norm or ('MCA' in pname_norm and 'MBA' not in pname_norm):
        return 'MCA'
    dept = (getattr(program, 'department', None) or '').strip()
    if dept and not any(tag in dept.upper() for tag in ('FYMCA', 'SYMCA', 'SEM-I', 'SEM-II', 'SEM-III', 'SEM-IV')):
        return dept
    return pname or dept or ''


def _enrich_criteria2_to_dict(d, rec):
    """Map canonical DB columns to Criteria 2 React field names."""
    tbl = rec.__tablename__
    if tbl == 'c211_enrolment':
        if getattr(rec, 'program_id', None):
            p = Program.query.get(rec.program_id)
            if p:
                d['programme_name'] = p.program_name
                d['programme_code'] = p.program_code
        d['seats_sanctioned'] = d.get('sanctioned_seats')
        d['students_admitted'] = d.get('admitted_students')
    elif tbl == 'c212_reservation':
        d['ear_sc'] = d.get('earmarked_sc')
        d['ear_st'] = d.get('earmarked_st')
        d['ear_obc'] = d.get('earmarked_obc')
        d['ear_gen'] = d.get('earmarked_gen')
        d['ear_others'] = d.get('earmarked_others')
        d['adm_sc'] = d.get('admitted_sc')
        d['adm_st'] = d.get('admitted_st')
        d['adm_obc'] = d.get('admitted_obc')
        d['adm_gen'] = d.get('admitted_gen')
        d['adm_others'] = d.get('admitted_others')
    elif tbl == 'c233_mentor_ratio':
        d['first'] = d.get('first_year_count')
        d['second'] = d.get('second_year_count')
        d['third'] = d.get('third_year_count')
        d['fourth'] = d.get('fourth_year_count')
        d['total'] = d.get('total_students')
    elif tbl == 'c241_teachers' and getattr(rec, 'teacher_id', None):
        d['teacher_id'] = rec.teacher_id
        d['department'] = getattr(rec, 'department', None) or ''
        t = Teacher.query.get(rec.teacher_id)
        if t:
            d['name'] = t.name
            d['pan'] = t.pan
            d['designation'] = t.designation
        d['year'] = d.get('year_of_appointment')
        d['nature'] = d.get('nature_of_appointment')
        d['experience'] = d.get('total_years_experience')
        ss = getattr(rec, 'is_still_serving', None)
        if ss is False:
            d['still_serving'] = 'No'
        else:
            d['still_serving'] = 'Yes'
    elif tbl == 'c242_teacher_phd' and getattr(rec, 'teacher_id', None):
        t = Teacher.query.get(rec.teacher_id)
        if t:
            d['teacher_name'] = t.name
        d['obtaining_year'] = d.get('year_of_obtaining')
    elif tbl == 'c23_outgoing_students':
        d['passing_year'] = d.get('year_of_passing')
        if getattr(rec, 'student_id', None):
            s = Student.query.get(rec.student_id)
            if s:
                d['student_name'] = s.name
                d['enrollment_number'] = s.enrollment_number
    elif tbl == 'c263_pass_percentage' and getattr(rec, 'program_id', None):
        p = Program.query.get(rec.program_id)
        if p:
            d['program_code'] = p.program_code
            d['program_name'] = _program_display_name(p)
        d['appeared_count'] = d.get('students_appeared')
        d['passed_count'] = d.get('students_passed')


def _parse_int(val, default=None):
    if val is None or val == '':
        return default
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _format_academic_year_display(val):
    """Show 2024 as 2024-25 for UI dropdowns."""
    if val is None or val == '':
        return None
    if isinstance(val, str) and '-' in val:
        return val.strip()
    try:
        y = int(val)
        return f"{y}-{str(y + 1)[-2:]}"
    except (TypeError, ValueError):
        return str(val)


def _year_start_from_offering_ui(val):
    """Map academic year dropdown values (e.g. 2026, '2024-25') to integer start year."""
    if val is None or val == '':
        return None
    if isinstance(val, int):
        return val
    s = str(val).strip()
    if s.isdigit():
        return int(s)
    # "2024-25" → 2024
    if "-" in s:
        left = s.split("-", 1)[0].strip()
        if left.isdigit():
            return int(left)
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _guess_course_code_from_type_label(label: str) -> str:
    """Bind experiential course-type labels (PBL, Major Project, …) to course_lookup rows."""
    if not label:
        return ""
    low = (label or "").lower()
    for c in Course.query.order_by(Course.id).all():
        cn = (c.course_name or "").lower()
        cc = (c.course_code or "").lower()
        if not cc:
            continue
        if "pbl" in low or "project based" in low:
            if "pbl" in cn or "project based" in cn or "pbl" in cc:
                return c.course_code
        if "major" in low:
            if "major" in cn or "major" in cc:
                return c.course_code
        if low in cn or cn in low:
            return c.course_code
    return ""


def _parse_int_or_sum(val, default=None):
    """Accepts '60+3' style entries for NAAC-style admitted counts."""
    if val is None or val == '':
        return default
    s = str(val).strip()
    if '+' in s:
        try:
            return sum(int(p.strip()) for p in s.split('+') if p.strip().isdigit())
        except ValueError:
            return default
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return default


def _resolve_student_for_outgoing(enrollment_number, name):
    enr = (enrollment_number or '').strip()
    nm = (name or '').strip()
    if not enr or not nm:
        return None
    s = Student.query.filter_by(enrollment_number=enr).first()
    if not s:
        s = Student(enrollment_number=enr, name=nm)
        db.session.add(s)
        db.session.flush()
    elif nm and s.name != nm:
        s.name = nm
    return s.id


def _resolve_student_by_name(name):
    """Match student_lookup row for Criterion 1.3.2 (tolerates extra spaces / case)."""
    nm = ' '.join((name or '').split()).strip()
    if not nm:
        return None
    stu = Student.query.filter(db.func.lower(Student.name) == nm.lower()).first()
    if stu:
        return stu
    return Student.query.filter(Student.name.ilike(nm)).first()


def _resolve_or_create_student_by_name(name):
    """Find student by name or create a minimal row (name-only uploads)."""
    nm = ' '.join((name or '').split()).strip()
    if not nm:
        return None
    stu = _resolve_student_by_name(nm)
    if stu:
        return stu
    slug = re.sub(r'[^a-z0-9]', '', nm.lower())[:20] or 'student'
    enr = f"NAAC-133-{slug}-{uuid.uuid4().hex[:6]}"[:100]
    while Student.query.filter_by(enrollment_number=enr).first():
        enr = f"NAAC-133-{slug}-{uuid.uuid4().hex[:8]}"[:100]
    stu = Student(name=nm, enrollment_number=enr)
    db.session.add(stu)
    db.session.flush()
    return stu


def _resolve_or_create_program(code, name=None):
    """Look up a programme by code; auto-create a minimal row if not found so FK constraints are satisfied."""
    code = (code or '').strip()
    if not code:
        return None
    p = Program.query.filter_by(program_code=code).first()
    if not p:
        p = Program(program_code=code, program_name=(name or code).strip() or code)
        db.session.add(p)
        db.session.flush()
    return p


def _resolve_or_create_course(code, name=None, program_id=None):
    """Look up course by unique code; auto-create for per-student codes in Criterion 1.3.2."""
    code = (code or '').strip()
    if not code:
        return None
    c = Course.query.filter_by(course_code=code).first()
    cname = (name or code).strip() or code
    if not c:
        c = Course(course_code=code, course_name=cname, program_id=program_id)
        db.session.add(c)
        db.session.flush()
    elif cname and (c.course_name or "").strip() != cname:
        c.course_name = cname
    if program_id and not c.program_id:
        c.program_id = program_id
    return c


def _finalize_criterion_create(criterion, model, data, db_kwargs):
    """Map frontend UI payloads to DB columns after the generic mapper."""
    tbl = model.__tablename__

    if tbl == 'c122_addon':
        # UI uses programName for "Name of Add-on/Certificate programme" → DB column `name`
        addon_title = (data.get('programName') or '').strip()
        if addon_title:
            db_kwargs['name'] = addon_title
        if data.get('duration') not in (None, ''):
            db_kwargs['duration'] = str(data['duration']).strip()
        yo = _year_start_from_offering_ui(data.get('yearOffering'))
        if yo is not None:
            db_kwargs['year_of_offering'] = yo
        if data.get('timesOffered') not in (None, ''):
            db_kwargs['times_offered'] = _parse_int(data.get('timesOffered'))
        if data.get('studentsEnrolled') not in (None, ''):
            db_kwargs['students_enrolled'] = _parse_int(data.get('studentsEnrolled'))
        if data.get('studentsCompleted') not in (None, ''):
            db_kwargs['students_completed'] = _parse_int(data.get('studentsCompleted'))

    if tbl == 'c132_experiential':
        lv = data.get('docLink') or data.get('pdfPath') or data.get('proof_links')
        if lv not in (None, ''):
            db_kwargs['proof_links'] = str(lv).strip()
        if data.get('year') not in (None, ''):
            ys = _year_start_from_offering_ui(data.get('year'))
            if ys is not None:
                db_kwargs['year_of_offering'] = ys
        cc = (data.get('courseCode') or '').strip()
        if cc and not db_kwargs.get('course_id'):
            cname = (data.get('courseType') or data.get('courseName') or cc).strip()
            c = _resolve_or_create_course(cc, cname, db_kwargs.get('program_id'))
            if c:
                db_kwargs['course_id'] = c.id

    if tbl == 'c121_cbcs':
        # Resolve program by programCode / programName
        pc = (data.get('programCode') or data.get('programme_code') or '').strip()
        pname = (data.get('programName') or data.get('programme_name') or pc or '').strip()
        if pc:
            p = _resolve_or_create_program(pc, pname)
            if p:
                db_kwargs['program_id'] = p.id
        db_kwargs['department'] = (data.get('department') or '').strip() or None
        db_kwargs['semester'] = (data.get('semester') or '').strip() or None
        db_kwargs['subject_code'] = (data.get('subjectCode') or '').strip() or None
        db_kwargs['subject_name'] = (data.get('subjectName') or '').strip() or None
        db_kwargs['year_intro'] = str(data['yearIntro']).strip() if data.get('yearIntro') not in (None, '') else None
        db_kwargs['cbcs_status'] = (data.get('cbcsStatus') or 'No').strip()
        db_kwargs['cbcs_year'] = str(data['cbcsYear']).strip() if data.get('cbcsYear') not in (None, '') else None

    if tbl == 'c211_enrolment':
        pc = data.get('programme_code') or data.get('programCode')
        pname = (data.get('programme_name') or data.get('programName') or pc or '').strip()
        p = _resolve_or_create_program(pc, pname)
        if p:
            db_kwargs['program_id'] = p.id
        if 'seats_sanctioned' in data and data.get('seats_sanctioned') not in (None, ''):
            db_kwargs['sanctioned_seats'] = _parse_int(data.get('seats_sanctioned'))
        if 'students_admitted' in data and data.get('students_admitted') not in (None, ''):
            db_kwargs['admitted_students'] = _parse_int_or_sum(data.get('students_admitted'))
        if not db_kwargs.get('academic_year'):
            db_kwargs['academic_year'] = session.get('academic_year') or ''

    if tbl == 'c212_reservation':
        if 'year' in data and data.get('year') is not None:
            db_kwargs['year'] = str(data['year']).strip()
        pairs = [
            ('ear_sc', 'earmarked_sc'), ('ear_st', 'earmarked_st'), ('ear_obc', 'earmarked_obc'),
            ('ear_gen', 'earmarked_gen'), ('ear_others', 'earmarked_others'),
            ('adm_sc', 'admitted_sc'), ('adm_st', 'admitted_st'), ('adm_obc', 'admitted_obc'),
            ('adm_gen', 'admitted_gen'), ('adm_others', 'admitted_others'),
        ]
        for ui_k, col_k in pairs:
            if ui_k in data and data[ui_k] not in (None, ''):
                db_kwargs[col_k] = _parse_int(data.get(ui_k), 0)

    if tbl == 'c23_outgoing_students':
        sid = _resolve_student_for_outgoing(data.get('enrollment_number'), data.get('student_name'))
        if sid:
            db_kwargs['student_id'] = sid
        if 'passing_year' in data:
            db_kwargs['year_of_passing'] = str(data['passing_year']).strip()

    if tbl == 'c233_mentor_ratio':
        db_kwargs['academic_year'] = (data.get('academic_year') or session.get('academic_year') or '').strip()
        db_kwargs['branch'] = (data.get('branch') or '').strip()
        pmap = [
            ('first', 'first_year_count'), ('second', 'second_year_count'),
            ('third', 'third_year_count'), ('fourth', 'fourth_year_count'),
        ]
        tot = 0
        for ui, col in pmap:
            v = _parse_int(data.get(ui), 0) or 0
            db_kwargs[col] = v
            tot += v
        db_kwargs['total_students'] = tot

    if tbl == 'c241_teachers':
        tid = None
        raw_tid = data.get('teacher_id')
        if raw_tid is not None and str(raw_tid).strip() != '':
            try:
                tid = int(raw_tid)
            except (TypeError, ValueError):
                tid = None
        if tid:
            t = Teacher.query.get(tid)
            if t:
                pan_in = (data.get('pan') or '').strip()
                if pan_in:
                    t.pan = pan_in
                des = (data.get('designation') or '').strip()
                if des:
                    t.designation = des
                db_kwargs['teacher_id'] = tid
        if not db_kwargs.get('teacher_id'):
            name = (data.get('name') or '').strip()
            if name:
                pan = (data.get('pan') or '').strip() or None
                t = Teacher.query.filter_by(pan=pan).first() if pan else None
                if not t:
                    t = Teacher.query.filter_by(name=name).first()
                if not t:
                    t = Teacher(name=name, pan=pan, designation=(data.get('designation') or '').strip() or None)
                    db.session.add(t)
                    db.session.flush()
                elif (data.get('designation') or '').strip():
                    t.designation = (data.get('designation') or '').strip()
                db_kwargs['teacher_id'] = t.id

        dept = (data.get('department') or '').strip()
        if not dept and session.get('user_id'):
            u = User.query.get(session['user_id'])
            if u and getattr(u, 'department', None):
                dept = (u.department or '').strip()
        if dept:
            db_kwargs['department'] = dept

        if 'year' in data and data.get('year') not in (None, ''):
            db_kwargs['year_of_appointment'] = _parse_int(data.get('year'))
        if data.get('nature') is not None:
            db_kwargs['nature_of_appointment'] = data['nature']
        if data.get('experience') not in (None, ''):
            try:
                db_kwargs['total_years_experience'] = float(data['experience'])
            except (TypeError, ValueError):
                pass

        if data.get('is_still_serving') is not None:
            sv = data['is_still_serving']
            serving = sv if isinstance(sv, bool) else str(sv).lower() in ('yes', 'true', '1')
            db_kwargs['is_still_serving'] = serving
            if serving:
                db_kwargs['last_year_of_service'] = None
            elif data.get('last_year_of_service') not in (None, ''):
                db_kwargs['last_year_of_service'] = _parse_int(data.get('last_year_of_service'))

    if tbl == 'c242_teacher_phd':
        tn = (data.get('teacher_name') or '').strip()
        if tn:
            t = Teacher.query.filter_by(name=tn).first()
            if not t:
                t = Teacher(name=tn)
                db.session.add(t)
                db.session.flush()
            db_kwargs['teacher_id'] = t.id
        if data.get('qualification') is not None:
            db_kwargs['qualification'] = data['qualification']
        if data.get('obtaining_year') not in (None, ''):
            db_kwargs['year_of_obtaining'] = _parse_int(data.get('obtaining_year'))
        if data.get('number_of_full_time_teachers') not in (None, ''):
            db_kwargs['number_of_full_time_teachers'] = _parse_int(data.get('number_of_full_time_teachers'))

    if tbl == 'c263_pass_percentage':
        pc = data.get('program_code') or data.get('programCode')
        pname = (data.get('program_name') or data.get('programName') or pc or '').strip()
        p = _resolve_or_create_program(pc, pname)
        if p:
            db_kwargs['program_id'] = p.id
        if 'year' in data and data.get('year') is not None:
            db_kwargs['year'] = str(data['year']).strip()
        if 'appeared_count' in data:
            db_kwargs['students_appeared'] = _parse_int(data.get('appeared_count'))
        if 'passed_count' in data:
            db_kwargs['students_passed'] = _parse_int(data.get('passed_count'))

    if tbl == 'c21_students_during_year' and data.get('enrollment_date'):
        try:
            p = str(data['enrollment_date'])[:10].split('-')
            db_kwargs['enrollment_date'] = date(int(p[0]), int(p[1]), int(p[2]))
        except Exception:
            db_kwargs.pop('enrollment_date', None)


def _apply_criterion_put(criterion, model, rec, data):
    tbl = model.__tablename__
    if tbl == 'c23_outgoing_students':
        if 'passing_year' in data:
            rec.year_of_passing = str(data['passing_year']).strip()
        if getattr(rec, 'student_id', None) and ('student_name' in data or 'enrollment_number' in data):
            s = Student.query.get(rec.student_id)
            if s:
                if 'student_name' in data:
                    s.name = data['student_name']
                if 'enrollment_number' in data:
                    s.enrollment_number = data['enrollment_number']
    if tbl == 'c241_teachers':
        new_tid = data.get('teacher_id')
        if new_tid is not None and str(new_tid).strip() != '':
            try:
                ntid = int(new_tid)
                if Teacher.query.get(ntid):
                    rec.teacher_id = ntid
            except (TypeError, ValueError):
                pass
        if getattr(rec, 'teacher_id', None):
            t = Teacher.query.get(rec.teacher_id)
            if t:
                if data.get('pan') is not None:
                    t.pan = data['pan'] or None
                if data.get('designation') is not None:
                    t.designation = data['designation']
        if data.get('department') is not None:
            rec.department = data['department'] or None
        if 'year' in data and data.get('year') not in (None, ''):
            rec.year_of_appointment = _parse_int(data.get('year'))
        if data.get('nature') is not None:
            rec.nature_of_appointment = data['nature']
        if data.get('experience') not in (None, ''):
            try:
                rec.total_years_experience = float(data['experience'])
            except (TypeError, ValueError):
                pass
        if data.get('is_still_serving') is not None:
            sv = data['is_still_serving']
            rec.is_still_serving = sv if isinstance(sv, bool) else str(sv).lower() in ('yes', 'true', '1')
            if rec.is_still_serving:
                rec.last_year_of_service = None
            elif 'last_year_of_service' in data:
                ly = data.get('last_year_of_service')
                rec.last_year_of_service = _parse_int(ly) if ly not in (None, '') else None
        elif 'last_year_of_service' in data:
            ly = data.get('last_year_of_service')
            rec.last_year_of_service = _parse_int(ly) if ly not in (None, '') else None
    if tbl == 'c263_pass_percentage':
        pc = data.get('program_code') or data.get('programCode')
        if pc:
            p = Program.query.filter_by(program_code=pc).first()
            if p:
                rec.program_id = p.id
        if 'year' in data:
            rec.year = str(data['year']).strip()
        if 'appeared_count' in data:
            rec.students_appeared = _parse_int(data.get('appeared_count'))
        if 'passed_count' in data:
            rec.students_passed = _parse_int(data.get('passed_count'))
    if tbl == 'c233_mentor_ratio':
        if 'branch' in data:
            rec.branch = (data.get('branch') or '').strip()
        if data.get('academic_year'):
            rec.academic_year = data['academic_year']
        pmap = [
            ('first', 'first_year_count'), ('second', 'second_year_count'),
            ('third', 'third_year_count'), ('fourth', 'fourth_year_count'),
        ]
        tot = 0
        for ui, col in pmap:
            if ui in data:
                v = _parse_int(data.get(ui), getattr(rec, col) or 0) or 0
                setattr(rec, col, v)
                tot += v
        rec.total_students = tot


def to_dict(rec):
    d = {}
    for col in rec.__table__.columns:
        val = getattr(rec, col.name)
        if isinstance(val, date): val = val.isoformat()
        if isinstance(val, Decimal): val = float(val)
        if isinstance(val, datetime): val = val.isoformat()
        d[col.name] = val
        
        # UI Aliasing
        if col.name == 'academic_year' or col.name == 'year_of_offering':
            d['year'] = _format_academic_year_display(val)
        if col.name == 'status_of_implementation':
            d['cbcsStatus'] = "Yes" if val else "No"
        if col.name == 'year_of_implementation':
            d['cbcsYear'] = val
        # proof_links → friendly alias for export
        if col.name == 'proof_links' and val:
            d['docLink'] = val
        if col.name == 'proof_document' and val:
            d['docLink'] = val
            d['documentPath'] = val
        # c121_cbcs camelCase aliases for React frontend
        if col.name == 'subject_code':
            d['subjectCode'] = val
        if col.name == 'subject_name':
            d['subjectName'] = val
        if col.name == 'year_intro':
            d['yearIntro'] = val
        if col.name == 'cbcs_status':
            if 'cbcsStatus' not in d:  # don't override the boolean alias above
                d['cbcsStatus'] = val
        if col.name == 'cbcs_year':
            if 'cbcsYear' not in d:
                d['cbcsYear'] = val
        # 1.1.3 academic bodies — camelCase for React
        if col.name == 'body_name':
            d['bodyName'] = val
        if col.name in DB_TO_UI:
            d[DB_TO_UI[col.name]] = val

    # Lookup related names for the UI tables
    if hasattr(rec, 'course_id') and rec.course_id:
        c = Course.query.get(rec.course_id)
        if c: d['courseCode'] = c.course_code; d['courseName'] = c.course_name
    if hasattr(rec, 'program_id') and rec.program_id:
        p = Program.query.get(rec.program_id)
        if p: d['programCode'] = p.program_code; d['programName'] = p.program_name
    if hasattr(rec, 'teacher_id') and rec.teacher_id:
        t = Teacher.query.get(rec.teacher_id)
        if t:
            d['teacherName'] = t.name
            d['teacher_id'] = t.id
            d['id_number_aadhar'] = t.aadhar_or_id
            d['email'] = t.email
            d['gender'] = t.gender
            d['designation'] = t.designation
            d['date_of_joining'] = t.joining_date.strftime('%d-%m-%Y') if isinstance(t.joining_date, (date, datetime)) and t.joining_date else ''
    if hasattr(rec, 'student_id') and rec.student_id:
        s = Student.query.get(rec.student_id)
        if s: d['enrollmentNumber'] = s.enrollment_number; d['studentName'] = s.name

    # Criterion 1.2.1 — programme name should be degree (MCA), not semester/FYMCA label
    if getattr(rec, '__tablename__', None) == 'c121_cbcs':
        if getattr(rec, 'program_id', None):
            p = Program.query.get(rec.program_id)
            if p:
                disp = _program_display_name(p)
                d['programName'] = disp or p.program_name
                d['programCode'] = p.program_code

    # Criterion 1.3.2 — course type label + explicit course_id already resolved above
    if getattr(rec, '__tablename__', None) == 'c132_experiential':
        d['courseType'] = None
        if getattr(rec, 'course_id', None):
            c = Course.query.get(rec.course_id)
            if c:
                d['courseCode'] = c.course_code
                d['courseName'] = c.course_name
                d['courseType'] = (c.course_name or '').strip() or None
                if not d['courseType']:
                    ctl = CourseTypeLookup.query.filter_by(course_code=c.course_code).first()
                    d['courseType'] = ctl.value if ctl else None
        if getattr(rec, 'program_id', None) and rec.program_id:
            p = Program.query.get(rec.program_id)
            if p:
                disp = _program_display_name(p)
                d['programName'] = disp or p.program_name
                d['programCode'] = p.program_code

    if getattr(rec, '__tablename__', None) == 'c133_projects':
        if getattr(rec, 'program_id', None):
            p = Program.query.get(rec.program_id)
            if p:
                disp = _program_display_name(p)
                d['programName'] = disp or p.program_name
                d['programCode'] = p.program_code
        if getattr(rec, 'student_id', None):
            s = Student.query.get(rec.student_id)
            if s:
                d['studentName'] = s.name
                d['enrollmentNumber'] = s.enrollment_number

    # Criterion 1.2.2 / 1.2.3 — addon title is stored in `name`; UI expects programName
    if getattr(rec, '__tablename__', None) == 'c122_addon':
        d['programName'] = getattr(rec, 'name', None)
        yo = getattr(rec, 'year_of_offering', None)
        d['yearOffering'] = yo
        to = getattr(rec, 'times_offered', None)
        d['timesOffered'] = to
        se = getattr(rec, 'students_enrolled', None)
        d['studentsEnrolled'] = se
        sc = getattr(rec, 'students_completed', None)
        d['studentsCompleted'] = sc
            
    # Resolve Usernames for auditing
    if hasattr(rec, 'created_by_id') and rec.created_by_id:
        u = User.query.get(rec.created_by_id)
        if u: d['createdBy'] = u.username
    if hasattr(rec, 'updated_by_id') and rec.updated_by_id:
        u = User.query.get(rec.updated_by_id)
        if u: d['updatedBy'] = u.username

    _enrich_criteria2_to_dict(d, rec)
    enrich_criteria346_to_dict(d, rec)
    return d


def serialize_c11_record(rec):
    """Normalize Criterion 1.1 response for frontend table rendering."""
    d = to_dict(rec)

    # Keep academic year explicit in both naming styles.
    d["academic_year"] = rec.academic_year
    d["academicYear"] = rec.academic_year

    if rec.program_id:
        p = Program.query.get(rec.program_id)
        if p:
            d["programCode"] = p.program_code
            if not d.get("programName"):
                d["programName"] = p.program_name

    # "Year of Introduction" should come from the course dimension.
    if rec.course_id:
        c = Course.query.get(rec.course_id)
        if c:
            d["courseCode"] = c.course_code
            d["courseName"] = c.course_name
            d["year"] = c.year_of_introduction

            # Semester label (e.g. FYMCA-SEM-I) overrides program name for NAAC export.
            scl = SemesterCourseLookup.query.filter_by(course_code=c.course_code).first()
            if scl:
                d["programName"] = scl.semester

    return d

@api_bp.route('/records/<criterion>', methods=['GET'])
def get_records(criterion):
    if criterion == '2_4_1_nature_options':
        recs = AppointmentTypeLookup.query.order_by(AppointmentTypeLookup.value).all()
        if not recs:
            for d in ["Regular", "Temporary", "Permanent"]:
                db.session.add(AppointmentTypeLookup(value=d))
            db.session.commit()
            recs = AppointmentTypeLookup.query.order_by(AppointmentTypeLookup.value).all()
        return jsonify([x.value for x in recs])
    if criterion == '2_4_2_qualification_options':
        recs = HighestDegreeLookup.query.order_by(HighestDegreeLookup.value).all()
        if not recs:
            for d in ["Ph.D.", "NET", "SET", "SLET"]:
                db.session.add(HighestDegreeLookup(value=d))
            db.session.commit()
            recs = HighestDegreeLookup.query.order_by(HighestDegreeLookup.value).all()
        return jsonify([x.value for x in recs])

    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify([])

    if criterion == '2_3_3_meta':
        m = C233MentorMeta.query.first()
        if not m:
            m = C233MentorMeta(mentor_count=1, academic_year=session.get('academic_year'))
            db.session.add(m)
            db.session.commit()
        return jsonify([to_dict(m)])
    
    if criterion == '1_1':
        records = C11Courses.query.all()
        return jsonify([serialize_c11_record(r) for r in records])

    c346_resp = get_records_c346(criterion, to_dict)
    if c346_resp is not None:
        return c346_resp
    
    return jsonify([to_dict(r) for r in model.query.all()])

# ---- Dedicated POST for Criterion 1.1 ----
@api_bp.route('/records/1_1', methods=['POST'])
def add_record_1_1():
    """
    Handles Criterion 1.1 form submission.
    UI sends: { department, programCode, programName (=semester), courseCode, courseName, year }
    Maps to C11Courses: { academic_year, program_id, course_id }
    """
    data = request.json or {}
    
    semester    = data.get('programName', '').strip()  # "FYMCA-SEM-I", etc.
    program_code= data.get('programCode', '').strip()
    course_code = data.get('courseCode', '').strip()
    intro_year  = str(data.get('year', '')).strip()
    academic_year = str(data.get('academicYear') or session.get('academic_year') or '').strip()
    department  = data.get('department', '').strip()

    # --- Validate required fields ---
    if not semester:
        return jsonify({"success": False, "error": "Semester (Program Name) is required"}), 400
    if not course_code:
        return jsonify({"success": False, "error": "Course Code is required"}), 400
    if not intro_year:
        return jsonify({"success": False, "error": "Year is required"}), 400
    if not academic_year and intro_year.isdigit():
        next_two = str(int(intro_year) + 1)[-2:]
        academic_year = f"{intro_year}-{next_two}"
    if not academic_year:
        return jsonify({"success": False, "error": "Academic Year is required"}), 400

    # --- 1. Resolve / auto-create Program ---
    prog = Program.query.filter_by(program_code=program_code).first() if program_code else None
    if not prog:
        # Create a minimal program entry so FK is satisfied
        prog = Program(
            program_code = program_code or semester,
            program_name = semester,
            department   = department
        )
        db.session.add(prog)
        db.session.flush()   # get prog.id without committing

    # --- 2. Resolve / auto-create Course (in the Course dim table) ---
    course = Course.query.filter_by(course_code=course_code).first()
    if not course:
        # Pull course name from the lookup table
        scl = SemesterCourseLookup.query.filter_by(
            semester=semester, course_code=course_code
        ).first()
        course_name = scl.course_name if scl else data.get('courseName', course_code)
        intro_year_int = int(intro_year) if intro_year.isdigit() else None
        course = Course(
            course_code=course_code,
            course_name=course_name,
            year_of_introduction=intro_year_int
        )
        db.session.add(course)
        db.session.flush()
    elif intro_year.isdigit() and not course.year_of_introduction:
        course.year_of_introduction = int(intro_year)

    # --- 3. Duplicate check: same year + program + course ---
    existing = C11Courses.query.filter_by(
        academic_year=academic_year,
        program_id=prog.id,
        course_id=course.id
    ).first()
    if existing:
        return jsonify({"success": False, "error": f"Course {course_code} for {semester} in {academic_year} already exists"}), 400

    # --- 4. Insert record ---
    try:
        new_rec = C11Courses(
            academic_year = academic_year,
            program_id    = prog.id,
            course_id     = course.id,
            proof_links   = data.get('proofLink'),
            created_by_id = session.get('user_id'),
            updated_by_id = session.get('user_id')
        )
        db.session.add(new_rec)
        db.session.commit()
        return jsonify({"success": True, "data": serialize_c11_record(new_rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/records/<criterion>', methods=['POST'])
def add_record(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify({"success": False, "error": "Criterion not found"}), 404

    data = request.json or {}

    # Mentor count singleton (2.3.3) — not a normal insert
    if criterion == '2_3_3_meta':
        m = C233MentorMeta.query.first()
        if not m:
            m = C233MentorMeta(mentor_count=1, academic_year=session.get('academic_year'))
            db.session.add(m)
            db.session.flush()
        cnt = _parse_int(data.get('mentor_count'), m.mentor_count)
        if cnt is not None:
            m.mentor_count = max(1, cnt)
        m.updated_by_id = session.get('user_id')
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(m)})

    db_kwargs = {}
    
    # Audit tracking
    user_id = session.get('user_id')
    db_kwargs['created_by_id'] = user_id
    db_kwargs['updated_by_id'] = user_id
    
    # Select All Optimization
    if data.get('select_all') is True:
        if 'programCode' in data:
            prog = Program.query.filter_by(program_code=data['programCode']).first()
            all_studs = Student.query.filter_by(program_id=prog.id).all() if prog else Student.query.all()
        else:
            all_studs = Student.query.all()
        data['student_list'] = ", ".join([s.name for s in all_studs])

    # 1. Map common foreign keys dynamically based on UI selections
    cc_in = (data.get('courseCode') or '').strip()
    pc_in = (data.get('programCode') or data.get('programme_code') or '').strip()
    prog_for_course = None
    if pc_in:
        prog_for_course = _resolve_or_create_program(pc_in, data.get('programName') or data.get('programme_name'))
    if cc_in and hasattr(model, 'course_id'):
        cname_in = (data.get('courseType') or data.get('courseName') or '').strip()
        c = _resolve_or_create_course(
            cc_in,
            cname_in,
            prog_for_course.id if prog_for_course else None,
        )
        if c:
            db_kwargs['course_id'] = c.id
    if pc_in and prog_for_course and hasattr(model, 'program_id'):
        db_kwargs['program_id'] = prog_for_course.id
    if data.get('programme_code'):
        p = _resolve_or_create_program(data['programme_code'], data.get('programme_name') or data.get('programName'))
        if p and hasattr(model, 'program_id'): db_kwargs['program_id'] = p.id
    raw_tid = data.get('teacher_id')
    if raw_tid not in (None, '') and hasattr(model, 'teacher_id'):
        try:
            tid = int(raw_tid)
            t = Teacher.query.get(tid)
            if t:
                db_kwargs['teacher_id'] = t.id
        except (TypeError, ValueError):
            pass
    if not db_kwargs.get('teacher_id') and 'teacherName' in data:
        tn = (data.get('teacherName') or '').strip()
        if tn:
            t = Teacher.query.filter_by(name=tn).first()
            if t and hasattr(model, 'teacher_id'):
                db_kwargs['teacher_id'] = t.id
    st_nm = (data.get('studentName') or data.get('student_name') or '').strip()
    if not st_nm and data.get('student_list'):
        st_nm = str(data.get('student_list')).split(',')[0].strip()
    if st_nm and hasattr(model, 'student_id'):
        stu = _resolve_student_by_name(st_nm)
        if not stu and getattr(model, '__tablename__', None) == 'c133_projects':
            stu = _resolve_or_create_student_by_name(st_nm)
        if stu:
            db_kwargs['student_id'] = stu.id
        
    # 2. Map payload keys to model columns
    for col in model.__table__.columns:
        col_name = col.name
        if col_name == 'id': continue
        
        # Direct match
        if col_name in data and data[col_name] is not None:
            if col_name == 'status_of_implementation':
                 db_kwargs[col_name] = str(data[col_name]).lower() == "yes"
            else:
                 db_kwargs[col_name] = data[col_name]
        else:
            # Common UI -> DB Alias mapping
            if col_name == 'academic_year' and 'year' in data:
                db_kwargs[col_name] = data['year']
            elif col_name == 'year_of_offering' and 'year' in data:
                ys = _year_start_from_offering_ui(data.get('year'))
                db_kwargs[col_name] = ys
            elif col_name == 'year_of_passing' and 'passing_year' in data:
                db_kwargs[col_name] = str(data['passing_year']).strip()
            elif col_name == 'year_of_passing' and 'year' in data:
                db_kwargs[col_name] = int(data['year']) if str(data['year']).isdigit() else data['year']
            elif col_name == 'year_of_implementation' and 'cbcsYear' in data:
                db_kwargs[col_name] = int(data['cbcsYear']) if str(data['cbcsYear']).isdigit() else None
            elif col_name == 'status_of_implementation' and 'cbcsStatus' in data:
                db_kwargs[col_name] = str(data['cbcsStatus']).lower() == "yes"
            elif col_name == 'year_of_obtaining' and 'obtaining_year' in data:
                db_kwargs[col_name] = _parse_int(data.get('obtaining_year'))
            elif col_name == 'proof_links' and col_name not in db_kwargs:
                # Accept docLink or pdfPath from frontend as proof_links
                link_val = data.get('docLink') or data.get('pdfPath') or data.get('proof_links')
                if link_val:
                    db_kwargs[col_name] = str(link_val).strip()
            elif col_name == 'body_name' and 'bodyName' in data:
                bv = data.get('bodyName')
                if bv not in (None, ''):
                    db_kwargs[col_name] = str(bv).strip()
            elif col_name == 'proof_document' and col_name not in db_kwargs:
                pv = (
                    data.get('proof_document') or data.get('docLink') or data.get('pdfPath')
                    or data.get('documentPath')
                )
                if pv not in (None, ''):
                    db_kwargs[col_name] = str(pv).strip()
            elif model.__tablename__ == 'c122_addon':
                # CamelCase ↔ snake_case not covered elsewhere
                if col_name == 'year_of_offering' and col_name not in db_kwargs:
                    ys = _year_start_from_offering_ui(data.get('yearOffering'))
                    if ys is not None:
                        db_kwargs[col_name] = ys
                elif col_name == 'times_offered' and col_name not in db_kwargs:
                    if data.get('timesOffered') not in (None, ''):
                        db_kwargs[col_name] = _parse_int(data.get('timesOffered'))
                elif col_name == 'students_enrolled' and col_name not in db_kwargs:
                    if data.get('studentsEnrolled') not in (None, ''):
                        db_kwargs[col_name] = _parse_int(data.get('studentsEnrolled'))
                elif col_name == 'students_completed' and col_name not in db_kwargs:
                    if data.get('studentsCompleted') not in (None, ''):
                        db_kwargs[col_name] = _parse_int(data.get('studentsCompleted'))
    if 'academic_year' not in db_kwargs and hasattr(model, 'academic_year'):
        db_kwargs['academic_year'] = session.get('academic_year')

    _finalize_criterion_create(criterion, model, data, db_kwargs)

    if criterion in CRITERIA_346_MODELS:
        apply_c346_foreign_keys(model, data, db_kwargs)
        apply_c346_column_mapping(model, data, db_kwargs)

    try:
        new_rec = model(**db_kwargs)
        db.session.add(new_rec)
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(new_rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/records/<criterion>/<int:id>', methods=['PUT'])
def update_record(criterion, id):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify({"success": False}), 404
    rec = model.query.get(id)
    if not rec: return jsonify({"success": False}), 404
    
    data = request.json or {}
    
    # Audit tracking
    rec.updated_by_id = session.get('user_id')
    
    # Select All Optimization
    if data.get('select_all') is True:
        if 'programCode' in data:
            prog = Program.query.filter_by(program_code=data['programCode']).first()
            all_studs = Student.query.filter_by(program_id=prog.id).all() if prog else Student.query.all()
        else:
            all_studs = Student.query.all()
        data['student_list'] = ", ".join([s.name for s in all_studs])

    pc_use = (data.get('programCode') or data.get('programme_code') or '').strip()
    p_upd = None
    if pc_use:
        p_upd = _resolve_or_create_program(pc_use, data.get('programName') or data.get('programme_name'))
        if p_upd and hasattr(model, 'program_id'):
            rec.program_id = p_upd.id
    if 'courseCode' in data:
        cc_in = (data['courseCode'] or '').strip()
        if cc_in and hasattr(model, 'course_id'):
            cname_in = (data.get('courseType') or data.get('courseName') or '').strip()
            c = _resolve_or_create_course(
                cc_in,
                cname_in,
                p_upd.id if p_upd else getattr(rec, 'program_id', None),
            )
            if c:
                rec.course_id = c.id
    if hasattr(model, 'teacher_id'):
        raw_tid = data.get('teacher_id')
        if raw_tid not in (None, ''):
            try:
                tid = int(raw_tid)
                t = Teacher.query.get(tid)
                if t:
                    rec.teacher_id = t.id
            except (TypeError, ValueError):
                pass
        if 'teacherName' in data:
            tn = (data.get('teacherName') or '').strip()
            if tn:
                t = Teacher.query.filter_by(name=tn).first()
                if t:
                    rec.teacher_id = t.id
    st_nm = (data.get('studentName') or data.get('student_name') or '').strip()
    if st_nm and hasattr(model, 'student_id'):
        if getattr(model, '__tablename__', None) == 'c133_projects':
            stu = _resolve_or_create_student_by_name(st_nm)
            if stu:
                rec.student_id = stu.id
        else:
            stu = Student.query.filter_by(name=st_nm).first()
            if stu:
                rec.student_id = stu.id

    # Criterion 1.1 custom behavior:
    # - keep academic_year from explicit academicYear/session, not from intro year.
    # - map "year" field to Course.year_of_introduction.
    if criterion == '1_1':
        if 'academicYear' in data and str(data['academicYear']).strip():
            rec.academic_year = str(data['academicYear']).strip()
        elif session.get('academic_year'):
            rec.academic_year = session.get('academic_year')

        if 'year' in data and rec.course_id:
            c = Course.query.get(rec.course_id)
            if c and str(data['year']).isdigit():
                c.year_of_introduction = int(data['year'])
        
    for col in model.__table__.columns:
        col_name = col.name
        if col_name == 'id': continue
        if col_name in data:
            if col_name == 'status_of_implementation':
                setattr(rec, col_name, str(data[col_name]).lower() == "yes")
            else:
                setattr(rec, col_name, data[col_name])
        else:
            if col_name == 'academic_year' and 'year' in data:
                setattr(rec, col_name, data['year'])
            elif col_name == 'year_of_offering' and 'year' in data:
                ys = _year_start_from_offering_ui(data.get('year'))
                if ys is not None:
                    setattr(rec, col_name, ys)
            elif col_name == 'year_of_passing' and 'passing_year' in data:
                setattr(rec, col_name, str(data['passing_year']).strip())
            elif col_name == 'year_of_implementation' and 'cbcsYear' in data:
                setattr(rec, col_name, int(data['cbcsYear']) if str(data['cbcsYear']).isdigit() else getattr(rec, col_name))
            elif col_name == 'status_of_implementation' and 'cbcsStatus' in data:
                setattr(rec, col_name, str(data['cbcsStatus']).lower() == "yes")
            elif col_name == 'nature_of_appointment' and 'nature' in data:
                setattr(rec, col_name, data['nature'])
            elif col_name == 'total_years_experience' and 'experience' in data:
                try:
                    setattr(rec, col_name, float(data['experience']))
                except (TypeError, ValueError):
                    pass
            elif col_name == 'proof_links':
                lv = data.get('docLink') or data.get('pdfPath') or data.get('proof_links')
                if lv not in (None, ''):
                    setattr(rec, col_name, str(lv).strip())
            elif col_name == 'year_of_appointment' and 'year' in data:
                setattr(rec, col_name, int(data['year']) if str(data['year']).isdigit() else getattr(rec, col_name))
            elif col_name == 'year_of_obtaining' and 'obtaining_year' in data:
                setattr(rec, col_name, _parse_int(data.get('obtaining_year')))
            elif col_name == 'body_name' and 'bodyName' in data:
                setattr(rec, col_name, str(data['bodyName'] or '').strip() or getattr(rec, col_name))
            elif col_name == 'proof_document':
                pv = (
                    data.get('proof_document') or data.get('docLink')
                    or data.get('pdfPath') or data.get('documentPath')
                )
                if pv not in (None, ''):
                    setattr(rec, col_name, str(pv).strip())
            elif getattr(rec, '__tablename__', None) == 'c122_addon':
                if col_name == 'name' and 'programName' in data:
                    nm = str(data['programName'] or '').strip()
                    if nm:
                        setattr(rec, col_name, nm)
                elif col_name == 'year_of_offering' and 'yearOffering' in data:
                    ys = _year_start_from_offering_ui(data.get('yearOffering'))
                    if ys is not None:
                        setattr(rec, col_name, ys)
                    elif data.get('yearOffering') in (None, ''):
                        setattr(rec, col_name, None)
                elif col_name == 'times_offered' and 'timesOffered' in data:
                    setattr(rec, col_name, _parse_int(data.get('timesOffered')))
                elif col_name == 'students_enrolled' and 'studentsEnrolled' in data:
                    setattr(rec, col_name, _parse_int(data.get('studentsEnrolled')))
                elif col_name == 'students_completed' and 'studentsCompleted' in data:
                    setattr(rec, col_name, _parse_int(data.get('studentsCompleted')))
                elif col_name == 'duration' and 'duration' in data:
                    setattr(rec, col_name, str(data['duration']).strip() or None)
    if criterion in CRITERIA_346_MODELS and criterion != '3_2_2':
        apply_c346_update(model, data, rec)
    db.session.commit()
    if criterion == '1_1':
        return jsonify({"success": True, "data": serialize_c11_record(rec)})
    return jsonify({"success": True, "data": to_dict(rec)})

@api_bp.route('/records/<criterion>/<int:id>', methods=['DELETE'])
def delete_record(criterion, id):
    model = CRITERIA_MODELS.get(criterion)
    if model:
        rec = model.query.get(id)
        if rec:
            db.session.delete(rec)
            db.session.commit()
    return jsonify({"success": True})

@api_bp.route('/records/<criterion>/bulk-delete', methods=['POST'])
def bulk_delete(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if model:
        ids = request.json.get('ids', [])
        model.query.filter(model.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
    return jsonify({"success": True})


# ============================================================
# EXCEL EXPORT (generic for any criterion)
# ============================================================

CRITERION_TITLES = {
    "1_1":   "1.1 Number of courses offered by the Institution across all programs during the year",
    "1_1_3": "1.1.3 Teachers participating in academic bodies",
    "1_2_1": "1.2.1 Number of Programmes in which Choice Based Credit System (CBCS)/ elective course system has been implemented",
    "1_2_2": "1.2.2 & 1.2.3 Add-on / Certificate programmes",
    "1_3_2": "1.3.2 Number of courses that include experiential learning through project work/field work/internship during the year",
    "1_3_3": "1.3.3 Number of students undertaking project work/field work/internships",
    "2_1":   "2.1 Number of students during the year",
    "2_2":   "2.2 Seats filled against reserved categories",
    "2_1_1": "2.1.1 Enrolment number",
    "2_1_2": "2.1.2 Seats earmarked and admitted (reserved categories)",
    "2_3":   "2.3 Outgoing students",
    "2_3_3": "2.3.3 Mentor–mentee ratio",
    "2_4_1": "2.4.1 Full-time teachers",
    "2_4_2": "2.4.2 Teachers with Ph.D. / NET / SET",
    "2_6_3": "2.6.3 Pass percentage of students",
    **CRITERION_346_TITLES,
}

# Columns to never include in NAAC exports
EXCLUDED_EXPORT_COLS = {
    "id", "course_id", "program_id", "teacher_id", "student_id",
    "created_by_id", "updated_by_id", "created_at", "updated_at",
    "createdBy", "updatedBy",
}

# Per-criterion explicit column whitelist (only these columns appear in data rows)
# Keys must match CRITERIA_MODELS keys. If a criterion is not listed here, all
# non-excluded columns are used.
CRITERION_COLUMN_WHITELIST = {
    "1_1_3": ["year", "teacherName", "bodyName"],
    "1_2_1": ["programCode", "programName", "yearIntro", "cbcsStatus", "cbcsYear"],
}

# Proof/link columns — shown ONCE as a merged row, never as data columns
EXCLUDED_LINK_COLS = {
    "docLink", "doc_link", "proof_links", "proof_document", "pdfPath", "link",
}

# Preferred column order for data rows (no link columns here)
_EXPORT_PREFERRED_ORDER = [
    "academic_year", "year", "year_of_offering", "yearOffering",
    "programCode", "programName", "programme_name", "programme_code",
    "department", "semester",
    "subjectCode", "subjectName", "yearIntro", "cbcsStatus", "cbcsYear",
    "courseCode", "courseName", "courseType",
    "teacherName", "bodyName", "body_name", "studentName", "enrollmentNumber", "enrollment_number",
    "student_name", "enrollment_year", "enrollment_date",
    "seats_sanctioned", "students_admitted",
    "ear_sc", "ear_st", "ear_obc", "ear_gen", "ear_others",
    "adm_sc", "adm_st", "adm_obc", "adm_gen", "adm_others",
    "appeared_count", "passed_count",
    "timesOffered", "duration", "studentsEnrolled", "studentsCompleted",
    "body_name",
]

# Human-readable header labels for column keys
_EXPORT_HEADERS = {
    "academic_year": "Academic Year", "year": "Academic Year",
    "year_of_offering": "Year of Offering", "yearOffering": "Year of Offering",
    "programCode": "Programme Code", "programName": "Programme Name",
    "programme_name": "Programme Name", "programme_code": "Programme Code",
    "department": "Department", "semester": "Semester / Type",
    "subjectCode": "Subject Code", "subjectName": "Subject Name",
    "yearIntro": "Year of Introduction",
    "cbcsStatus": "Status of implementation of CBCS / elective course system (Yes/No)",
    "cbcsYear": "Year of implementation of CBCS / elective course system",
    "courseCode": "Course Code", "courseName": "Course Name", "courseType": "Course Type",
    "teacherName": "Teacher Name", "studentName": "Student Name",
    "enrollmentNumber": "Enrollment Number", "enrollment_number": "Enrollment Number",
    "student_name": "Student Name", "enrollment_year": "Enrollment Year",
    "enrollment_date": "Enrollment Date",
    "seats_sanctioned": "Seats Sanctioned", "students_admitted": "Students Admitted",
    "ear_sc": "Earmarked SC", "ear_st": "Earmarked ST", "ear_obc": "Earmarked OBC",
    "ear_gen": "Earmarked General", "ear_others": "Earmarked Others",
    "adm_sc": "Admitted SC", "adm_st": "Admitted ST", "adm_obc": "Admitted OBC",
    "adm_gen": "Admitted General", "adm_others": "Admitted Others",
    "appeared_count": "Students Appeared", "passed_count": "Students Passed",
    "timesOffered": "Times Offered", "duration": "Duration",
    "studentsEnrolled": "Students Enrolled", "studentsCompleted": "Students Completed",
    "body_name": "Academic Body Name", "proof_document": "Proof / Document Link",
    "doc_link": "Document Link", "docLink": "Document Link",
    "proof_links": "Proof / Document Link", "pdfPath": "Document Link",
    "student_list": "Student Names",
}

def _export_header(col: str) -> str:
    return _EXPORT_HEADERS.get(col) or col.replace("_", " ").title()


from models.models import ProofLink11, ProofLink121, ProofLink122_123, CriterionProofLink

# Legacy specific-model map (kept for backwards compat with existing DB rows)
_LEGACY_PROOF_MODELS = {
    "1_1":   ProofLink11,
    "1_2_1": ProofLink121,
    "1_2_2": ProofLink122_123,
}

def _get_proof_link_value(criterion: str) -> str | None:
    """Return saved proof link for any criterion (legacy or generic table)."""
    if criterion in _LEGACY_PROOF_MODELS:
        pl = _LEGACY_PROOF_MODELS[criterion].query.first()
        return pl.link if pl else None
    pl = CriterionProofLink.query.filter_by(criterion_key=criterion).first()
    return pl.link if pl else None

def _save_proof_link_value(criterion: str, link: str) -> None:
    """Upsert proof link for any criterion."""
    if criterion in _LEGACY_PROOF_MODELS:
        pm = _LEGACY_PROOF_MODELS[criterion]
        pl = pm.query.first()
        if not pl:
            pl = pm(link=link)
            db.session.add(pl)
        else:
            pl.link = link
    else:
        pl = CriterionProofLink.query.filter_by(criterion_key=criterion).first()
        if not pl:
            pl = CriterionProofLink(criterion_key=criterion, link=link)
            db.session.add(pl)
        else:
            pl.link = link
    db.session.commit()

# Map criterion key → proof link for Excel export (uses the helper above now)
_GLOBAL_PROOF_MODELS = None  # no longer needed — use _get_proof_link_value directly

EXPORT_PROOF_COL_HEADER = "Document / Proof Link"

_C132_EXPORT_HEADERS = [
    "Program name",
    "Program code",
    "Name of the Course that include experiential learning through project work/field work/internship",
    "Course code",
    "Year of offering",
    "Name of the student studied course on experiential learning through project work/field work/internship",
    EXPORT_PROOF_COL_HEADER,
]


_C11_EXPORT_HEADERS = [
    "Program code",
    "Program Name",
    "Course code",
    "Course Name",
    "Year of introduction",
]


def _apply_merged_proof_column(ws, *, first_data_row, last_data_row, doc_col_idx,
                               global_proof_link, thin_border, center_align):
    """Write one combined proof link merged vertically across all data rows."""
    if last_data_row < first_data_row:
        return
    for ri in range(first_data_row, last_data_row + 1):
        ws.cell(row=ri, column=doc_col_idx, value="")
    if last_data_row > first_data_row:
        ws.merge_cells(
            start_row=first_data_row, start_column=doc_col_idx,
            end_row=last_data_row, end_column=doc_col_idx,
        )
    link_cell = ws.cell(row=first_data_row, column=doc_col_idx)
    raw_link = (global_proof_link or "").strip()
    if raw_link:
        full_url = _proof_link_to_excel_full_url(raw_link)
        link_cell.value = full_url
        link_cell.hyperlink = full_url
        link_cell.font = Font(name='Calibri', size=9, color='0563C1', underline='single')
    else:
        link_cell.value = "(No proof link saved yet)"
        link_cell.font = Font(name='Calibri', size=9)
    link_cell.alignment = center_align
    link_cell.border = thin_border


def _export_excel_1_1(rows, global_proof_link):
    """NAAC template layout for Criterion 1.1 — five data columns + merged proof link."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Criteria 1.1"

    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, name='Calibri', size=11)
    header_font = Font(bold=True, name='Calibri', size=10)
    data_font = Font(name='Calibri', size=10)
    hdr_fill = PatternFill('solid', start_color='D9D9D9')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

    doc_col = len(_C11_EXPORT_HEADERS) + 1
    last_col = get_column_letter(doc_col)

    ws.merge_cells(f'A1:{last_col}1')
    t = ws['A1']
    t.value = CRITERION_TITLES["1_1"]
    t.font = title_font
    t.alignment = left_wrap
    t.border = thin_border
    ws.row_dimensions[1].height = 36

    for ci, head in enumerate(_C11_EXPORT_HEADERS, start=1):
        c = ws.cell(row=2, column=ci, value=head)
        c.font = header_font
        c.alignment = center
        c.border = thin_border
        c.fill = hdr_fill
    ws.cell(row=2, column=doc_col, value=EXPORT_PROOF_COL_HEADER).font = header_font
    ws.cell(row=2, column=doc_col).alignment = center
    ws.cell(row=2, column=doc_col).border = thin_border
    ws.cell(row=2, column=doc_col).fill = hdr_fill
    ws.row_dimensions[2].height = 32

    first_data_row = 3
    last_data_row = first_data_row + len(rows) - 1 if rows else first_data_row - 1
    for ri, row in enumerate(rows, start=first_data_row):
        values = [
            row.get('programCode') or '',
            row.get('programName') or '',
            row.get('courseCode') or '',
            row.get('courseName') or '',
            row.get('year') or '',
        ]
        for ci, val in enumerate(values, start=1):
            c = ws.cell(row=ri, column=ci, value="" if val is None else val)
            c.font = data_font
            c.alignment = left_wrap
            c.border = thin_border
        ws.cell(row=ri, column=doc_col, value="")
        for ci in range(1, doc_col + 1):
            ws.cell(row=ri, column=ci).border = thin_border

    _apply_merged_proof_column(
        ws,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        doc_col_idx=doc_col,
        global_proof_link=global_proof_link,
        thin_border=thin_border,
        center_align=center,
    )

    widths = [16, 28, 16, 40, 18, 40]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Criteria_1_1_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _export_excel_1_3_2(rows, global_proof_link):
    """NAAC template layout for Criterion 1.3.2 (7 columns, no Sr. No.)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Criteria 1.3.2"

    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, name='Calibri', size=11)
    subtitle_font = Font(bold=True, name='Calibri', size=10, color='FF0000')
    header_font = Font(bold=True, name='Calibri', size=10)
    data_font = Font(name='Calibri', size=10)
    hdr_fill = PatternFill('solid', start_color='D9D9D9')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

    last_col = get_column_letter(len(_C132_EXPORT_HEADERS))

    ws.merge_cells(f'A1:{last_col}1')
    t = ws['A1']
    t.value = CRITERION_TITLES["1_3_2"]
    t.font = title_font
    t.alignment = left_wrap
    t.border = thin_border
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f'A2:{last_col}2')
    sub = ws['A2']
    sub.value = "Consider Project Stage 1, Internship, Field work if any"
    sub.font = subtitle_font
    sub.alignment = left_wrap
    sub.border = thin_border
    ws.row_dimensions[2].height = 22

    for ci, head in enumerate(_C132_EXPORT_HEADERS, start=1):
        c = ws.cell(row=3, column=ci, value=head)
        c.font = header_font
        c.alignment = center
        c.border = thin_border
        c.fill = hdr_fill
    ws.row_dimensions[3].height = 48

    doc_col = 7
    first_data_row = 4
    last_data_row = first_data_row + len(rows) - 1 if rows else first_data_row - 1

    for ri, row in enumerate(rows, start=first_data_row):
        course_label = (row.get('courseType') or row.get('courseName') or '').strip()
        values = [
            row.get('programName') or '',
            row.get('programCode') or '',
            course_label,
            row.get('courseCode') or '',
            row.get('year') or '',
            row.get('studentName') or '',
        ]
        for ci, val in enumerate(values, start=1):
            c = ws.cell(row=ri, column=ci, value="" if val is None else val)
            c.font = data_font
            c.alignment = left_wrap
            c.border = thin_border
        ws.cell(row=ri, column=doc_col, value="")
        for ci in range(1, doc_col + 1):
            ws.cell(row=ri, column=ci).border = thin_border

    _apply_merged_proof_column(
        ws,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        doc_col_idx=doc_col,
        global_proof_link=global_proof_link,
        thin_border=thin_border,
        center_align=center,
    )

    widths = [18, 14, 42, 14, 16, 42, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Criteria_1_3_2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


_C133_EXPORT_HEADERS = [
    "Program name",
    "Program code",
    "Name of students undertaking project work/field work/internships",
    EXPORT_PROOF_COL_HEADER,
]


def _export_excel_1_3_3(rows, global_proof_link):
    """NAAC template layout for Criterion 1.3.3."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Criteria 1.3.3"

    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, name='Calibri', size=11)
    header_font = Font(bold=True, name='Calibri', size=10)
    data_font = Font(name='Calibri', size=10)
    hdr_fill = PatternFill('solid', start_color='D9D9D9')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

    last_col = get_column_letter(len(_C133_EXPORT_HEADERS))

    ws.merge_cells(f'A1:{last_col}1')
    t = ws['A1']
    t.value = CRITERION_TITLES["1_3_3"]
    t.font = title_font
    t.alignment = left_wrap
    t.border = thin_border
    ws.row_dimensions[1].height = 36

    for ci, head in enumerate(_C133_EXPORT_HEADERS, start=1):
        c = ws.cell(row=2, column=ci, value=head)
        c.font = header_font
        c.alignment = center
        c.border = thin_border
        c.fill = hdr_fill
    ws.row_dimensions[2].height = 40

    doc_col = 4
    first_data_row = 3
    last_data_row = first_data_row + len(rows) - 1 if rows else first_data_row - 1

    for ri, row in enumerate(rows, start=first_data_row):
        values = [
            row.get('programName') or '',
            row.get('programCode') or '',
            row.get('studentName') or '',
        ]
        for ci, val in enumerate(values, start=1):
            c = ws.cell(row=ri, column=ci, value="" if val is None else val)
            c.font = data_font
            c.alignment = left_wrap
            c.border = thin_border
        ws.cell(row=ri, column=doc_col, value="")
        for ci in range(1, doc_col + 1):
            ws.cell(row=ri, column=ci).border = thin_border

    _apply_merged_proof_column(
        ws,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        doc_col_idx=doc_col,
        global_proof_link=global_proof_link,
        thin_border=thin_border,
        center_align=center,
    )

    for i, w in enumerate([18, 14, 42, 36], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Criteria_1_3_3_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@api_bp.route('/export-excel/<criterion>', methods=['GET'])
def export_excel(criterion):
    """Plain NAAC-ready Excel export — no colours, no audit columns, proof links included."""
    model = CRITERIA_MODELS.get(criterion)
    if not model:
        return jsonify({"success": False, "error": f"Unknown criterion: {criterion}"}), 404

    records = model.query.all()
    rows = [to_dict(r) for r in records]

    # Fetch the single global proof link (works for any criterion key)
    global_proof_link = _get_proof_link_value(criterion)
    # Fallback: pick the first non-empty per-record proof link
    if not global_proof_link:
        for row in rows:
            lnk = row.get('docLink') or row.get('proof_links') or row.get('proof_document')
            if lnk:
                global_proof_link = lnk
                break

    if criterion == "1_1":
        records = C11Courses.query.all()
        rows = [serialize_c11_record(r) for r in records]
        return _export_excel_1_1(rows, global_proof_link)

    if criterion == "1_3_2":
        return _export_excel_1_3_2(rows, global_proof_link)

    if criterion == "1_3_3":
        return _export_excel_1_3_3(rows, global_proof_link)

    # Build data columns — exclude ALL link/proof-related keys (shown in a single merged row)
    _alias_pairs = [
        ("academic_year", "year"),
        ("programCode", "program_code"), ("programName", "program_name"),
        ("subjectCode", "subject_code"), ("subjectName", "subject_name"),
        ("yearIntro", "year_intro"), ("cbcsStatus", "cbcs_status"), ("cbcsYear", "cbcs_year"),
        ("courseCode", "course_code"), ("courseName", "course_name"),
        ("teacherName", "teacher_name"), ("studentName", "student_name"),
        ("enrollmentNumber", "enrollment_number"),
        ("yearOffering", "year_of_offering"),
        ("timesOffered", "times_offered"), ("studentsEnrolled", "students_enrolled"),
        ("studentsCompleted", "students_completed"),
        ("appeared_count", "students_appeared"), ("passed_count", "students_passed"),
    ]

    _all_excluded = EXCLUDED_EXPORT_COLS | EXCLUDED_LINK_COLS

    # Use explicit whitelist if defined for this criterion
    whitelist = CRITERION_COLUMN_WHITELIST.get(criterion)
    if whitelist:
        columns = whitelist
    elif rows:
        all_keys: set = set()
        for r in rows:
            all_keys.update(r.keys())
        all_keys -= _all_excluded  # strip audit + ALL link-related keys

        for keep, drop in _alias_pairs:
            if keep in all_keys and drop in all_keys:
                all_keys.discard(drop)

        columns = [k for k in _EXPORT_PREFERRED_ORDER if k in all_keys]
        for k in sorted(all_keys):
            if k not in columns:
                columns.append(k)
    else:
        columns = [c.name for c in model.__table__.columns
                   if c.name not in _all_excluded]

    title_text = CRITERION_TITLES.get(criterion, f"Criteria {criterion.replace('_', '.')}")

    # Column layout:
    #   col 1        = Sr. No.
    #   col 2..N     = data columns
    #   col N+1      = Document / Proof Link  (header row 2; merged vertically across all data rows)
    num_data_cols = len(columns)
    total_cols    = num_data_cols + 2          # +1 Sr.No  +1 doc-link
    last_col_letter = get_column_letter(total_cols)
    doc_col_idx   = total_cols                 # last column index (1-based)
    doc_col_letter = get_column_letter(doc_col_idx)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Criteria {criterion.replace('_', '.')}"

    # ── Style definitions ─────────────────────────────────────────────────────
    title_font  = Font(bold=True,  name='Calibri', size=12)
    header_font = Font(bold=True,  name='Calibri', size=10)
    data_font   = Font(bold=False, name='Calibri', size=10)
    link_val_font = Font(bold=False, name='Calibri', size=9)

    center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_ctr = Alignment(horizontal='center', vertical='top',    wrap_text=True)
    data_al = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', start_color='D9D9D9')   # grey header fill

    # ── Row 1 : Criteria title — merged across ALL columns ───────────────────
    ws.merge_cells(f'A1:{last_col_letter}1')
    c = ws['A1']
    c.value     = title_text
    c.font      = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Row 2 : Column headers ────────────────────────────────────────────────
    # Sr. No.
    ws.cell(row=2, column=1, value="Sr.\nNo.")
    # Data column headers
    for ci, col in enumerate(columns, start=2):
        ws.cell(row=2, column=ci, value=_export_header(col))
    # Document / Proof Link header (last column)
    ws.cell(row=2, column=doc_col_idx, value=EXPORT_PROOF_COL_HEADER)

    for ci in range(1, total_cols + 1):
        c = ws.cell(row=2, column=ci)
        c.font      = header_font
        c.alignment = center
        c.border    = thin_border
        c.fill      = hdr_fill
    ws.row_dimensions[2].height = 36

    # ── Rows 3+ : Data rows ───────────────────────────────────────────────────
    first_data_row = 3
    last_data_row  = max(first_data_row, first_data_row + len(rows) - 1)

    for ri, row in enumerate(rows, start=first_data_row):
        ws.cell(row=ri, column=1, value=ri - first_data_row + 1)   # Sr. No.
        for ci, col in enumerate(columns, start=2):
            val = row.get(col)
            ws.cell(row=ri, column=ci, value="" if val is None else val)
        # Document link column — value written only on first row; all cells get border
        ws.cell(row=ri, column=doc_col_idx, value="")
        for ci in range(1, total_cols + 1):
            c = ws.cell(row=ri, column=ci)
            c.font      = data_font
            c.alignment = data_al
            c.border    = thin_border

    # ── Merge document-link column across all data rows & fill with link ──────
    _apply_merged_proof_column(
        ws,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        doc_col_idx=doc_col_idx,
        global_proof_link=global_proof_link,
        thin_border=thin_border,
        center_align=center,
    )

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 6
    for ci, col in enumerate(columns, start=2):
        col_letter = get_column_letter(ci)
        h = _export_header(col).lower()
        if any(x in h for x in ("name", "programme", "course", "student", "body", "list")):
            ws.column_dimensions[col_letter].width = 34
        elif any(x in h for x in ("year", "date", "code", "status", "semester")):
            ws.column_dimensions[col_letter].width = 18
        elif any(x in h for x in ("enrolled", "admitted", "sanctioned", "appeared", "passed", "offered")):
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions[doc_col_letter].width = 40   # document link column

    ws.freeze_panes = 'B3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Criteria_{criterion}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============================================================
# CRITERIA 2.1 — Parse student list PDF (enrollment no. + name)
# ============================================================

# Enrollment number must contain at least one digit to avoid matching plain names like "John"
_C21_ENROLL_TOKEN = re.compile(r"^(?=[A-Za-z0-9/_-]*\d)[A-Za-z0-9/_-]{4,28}$")
_C21_MAX_STUDENTS = 800
_C21_MAX_PDF_BYTES = 15 * 1024 * 1024


def _c21_clean_student_name(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip()).strip(".,;: ")
    return s


def _c21_is_enrollment(token: str) -> bool:
    """Return True only if token looks like an enrollment/PRN number (must contain a digit)."""
    return bool(_C21_ENROLL_TOKEN.match(token))


def _c21_is_name_fragment(token: str) -> bool:
    """Return True if token looks like part of a person's name (has letters, no digits)."""
    return bool(re.search(r"[A-Za-z]", token)) and not bool(re.search(r"\d", token))


def _c21_parse_student_line(line: str) -> dict | None:
    """Infer one student from a single text line (tabular PDF extraction).

    Handles common Indian university PDF formats:
      Format A:  SR  ENROLLMENT  NAME ...         (cols split by 2+ spaces)
      Format B:  SR  NAME  ENROLLMENT             (enrollment at end)
      Format C:  ENROLLMENT  NAME                 (two columns, no SR)
      Format D:  SR. NAME ENROLLMENT              (single space, enrollment is last token with digit)
      Format E:  NAME, ENROLLMENT  or  ENROLLMENT, NAME
    """
    if not line or len(line.strip()) < 6:
        return None
    line = line.strip()
    compact = re.sub(r"\s+", " ", line)

    first = compact.split()[0].lower().rstrip(".") if compact.split() else ""
    header_starts = (
        "sr", "sl", "s.no", "sn", "enrollment", "enrolment", "reg", "register",
        "name", "student", "programme", "program", "department", "course",
        "signature", "remark", "page", "annexure", "naac", "no",
    )
    if first in header_starts and len(compact) < 80:
        return None

    # --- Strategy 1: columns clearly separated by 2+ spaces ---
    cols = [c.strip() for c in re.split(r"\s{2,}", line) if c.strip()]
    if len(cols) >= 3:
        # Format A: SR | ENROLLMENT | NAME...
        if cols[0].replace(".", "").isdigit() and _c21_is_enrollment(cols[1]):
            name = _c21_clean_student_name(" ".join(cols[2:]))
            if len(name) >= 2 and re.search(r"[A-Za-z]", name):
                return {"enrollment_number": cols[1], "student_name": name}
        # Format B: SR | NAME... | ENROLLMENT  (enrollment is last col)
        if cols[0].replace(".", "").isdigit() and _c21_is_enrollment(cols[-1]):
            name = _c21_clean_student_name(" ".join(cols[1:-1]))
            if len(name) >= 2 and re.search(r"[A-Za-z]", name):
                return {"enrollment_number": cols[-1], "student_name": name}
        # Format A without SR: ENROLLMENT | NAME...
        if _c21_is_enrollment(cols[0]):
            name = _c21_clean_student_name(" ".join(cols[1:]))
            if len(name) >= 2 and re.search(r"[A-Za-z]", name):
                return {"enrollment_number": cols[0], "student_name": name}

    if len(cols) == 2:
        a, b = cols[0], cols[1]
        if _c21_is_enrollment(a) and re.search(r"[A-Za-z]", b):
            return {"enrollment_number": a, "student_name": _c21_clean_student_name(b)}
        if _c21_is_enrollment(b) and re.search(r"[A-Za-z]", a):
            return {"enrollment_number": b, "student_name": _c21_clean_student_name(a)}

    # --- Strategy 2: single-space split ---
    parts = compact.split()
    if len(parts) >= 3:
        # Determine if the first token is a serial number
        i = 0
        if parts[0].rstrip(").").isdigit() or (parts[0].endswith(".") and parts[0][:-1].isdigit()):
            i = 1

        # Format A-single: SR ENROLLMENT NAME...
        if i < len(parts) and _c21_is_enrollment(parts[i]):
            name = _c21_clean_student_name(" ".join(parts[i + 1:]))
            if len(name) >= 2 and re.search(r"[A-Za-z]", name):
                return {"enrollment_number": parts[i], "student_name": name}

        # Format B-single: SR NAME... ENROLLMENT  (enrollment is last token, contains digit)
        if _c21_is_enrollment(parts[-1]):
            name_parts = parts[i:-1]
            name = _c21_clean_student_name(" ".join(name_parts))
            if len(name) >= 2 and re.search(r"[A-Za-z]", name):
                return {"enrollment_number": parts[-1], "student_name": name}

    # --- Strategy 3: comma-separated ---
    if "," in compact:
        left, _, right = compact.partition(",")
        left, right = left.strip(), right.strip()
        if _c21_is_enrollment(left) and re.search(r"[A-Za-z]", right):
            return {"enrollment_number": left, "student_name": _c21_clean_student_name(right)}
        if _c21_is_enrollment(right) and re.search(r"[A-Za-z]", left):
            return {"enrollment_number": right, "student_name": _c21_clean_student_name(left)}

    # --- Strategy 4: regex fallback (enrollment must contain a digit) ---
    m = re.match(
        r"^\s*(?:\d{1,4}[.)]\s*)?(?:\d+\s+)?([A-Za-z0-9/_-]{4,28})\s+(.{2,120})$",
        compact,
    )
    if m:
        enr, name = m.group(1), _c21_clean_student_name(m.group(2))
        if _c21_is_enrollment(enr) and re.search(r"[A-Za-z]", name):
            return {"enrollment_number": enr, "student_name": name}

    return None


_C21_PURE_NAME_RE = re.compile(r"^[A-Za-z][A-Za-z\s.''\-]{2,80}$")
_C21_HEADER_LINES = {
    "name of students", "student enrollment number", "name", "enrollment number",
    "enrolment number", "sr no", "sr.", "sl no", "s.no", "roll no", "roll number",
}


def _c21_two_column_fallback(lines: list) -> list:
    """
    Handle PDFs where pypdf extracts a two-column layout as two separate text blocks:
    all names first, then all enrollment numbers (or vice versa).
    Zip them by position.
    """
    names = []
    enrollments = []
    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        if line.lower().strip(".:") in _C21_HEADER_LINES:
            continue
        if _c21_is_enrollment(line):
            enrollments.append(line)
        elif _C21_PURE_NAME_RE.match(line) and not re.search(r"\d", line):
            names.append(_c21_clean_student_name(line))

    if not names or not enrollments:
        return []
    # Only use this strategy when counts are close (same column length ± 10%)
    ratio = min(len(names), len(enrollments)) / max(len(names), len(enrollments))
    if ratio < 0.5:
        return []
    out = []
    seen = set()
    for i, name in enumerate(names):
        if i >= len(enrollments):
            break
        key = enrollments[i].upper()
        if key in seen:
            continue
        seen.add(key)
        out.append({"enrollment_number": enrollments[i], "student_name": name})
    return out


def _c21_extract_students_from_pdf_bytes(data: bytes) -> tuple:
    """Returns (students_list, raw_lines_list) for debugging."""
    reader = PdfReader(io.BytesIO(data))
    chunks = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            chunks.append(t)
    full = "\n".join(chunks)
    lines = []
    for block in full.replace("\r", "").split("\n"):
        for seg in block.split("\x0c"):
            seg = seg.strip()
            if seg:
                lines.append(seg)

    # Strategy 1: line-by-line parsing (works when each row has both fields on one line)
    seen = set()
    out = []
    for raw in lines:
        rec = _c21_parse_student_line(raw)
        if not rec:
            continue
        key = rec["enrollment_number"].upper()
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
        if len(out) >= _C21_MAX_STUDENTS:
            break

    # Strategy 2: two-column block fallback (pypdf often extracts all names first, then all
    # enrollment numbers as separate text blocks — zip them by position)
    fallback = _c21_two_column_fallback(lines)

    # Use whichever strategy found more students
    if len(fallback) > len(out):
        out = fallback

    return out, lines


@api_bp.route('/criteria2/2_1/parse-student-list-pdf', methods=['POST'])
def parse_criterion_21_student_list_pdf():
    """Upload a PDF with selectable text; returns enrollment_number + student_name rows."""
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded", "students": []}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"success": False, "error": "Empty filename", "students": []}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext != '.pdf':
        return jsonify({"success": False, "error": "Only PDF files are supported", "students": []}), 400
    data = f.read()
    if not data:
        return jsonify({"success": False, "error": "Empty file", "students": []}), 400
    if len(data) > _C21_MAX_PDF_BYTES:
        return jsonify({"success": False, "error": "File too large (max 15 MB)", "students": []}), 400
    debug_mode = request.args.get("debug", "").lower() in ("1", "true", "yes")
    try:
        students, raw_lines = _c21_extract_students_from_pdf_bytes(data)
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read PDF: {str(e)}", "students": []}), 400

    if not students:
        resp = {
            "success": False,
            "error": (
                "No student rows detected. The PDF must have selectable text (not a scanned/image-only file). "
                "Ensure each row has an enrollment number (containing digits) and a student name. "
                "Use the 'Preview raw PDF text' button to see what text was extracted."
            ),
            "students": [],
        }
        if debug_mode:
            resp["raw_lines"] = raw_lines[:200]
        return jsonify(resp)

    resp = {"success": True, "students": students, "count": len(students)}
    if debug_mode:
        resp["raw_lines"] = raw_lines[:200]
    return jsonify(resp)


def _c21_clean_cell_str(val) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    s = str(val).strip()
    if s.lower() in ("nan", "nat", "none"):
        return ""
    if re.fullmatch(r"\d+\.0", s):
        return s[:-2]
    return s


def _c21_parse_date_cell(val) -> date | None:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("nan", "nat", "none"):
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(s, dayfirst=False, errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
    except Exception:
        pass
    try:
        parsed = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
    except Exception:
        pass
    return None


def _c21_compact_header(c) -> str:
    return re.sub(r"[\s._\-/]+", "", _c23_normalize_header(c))


def _c21_pick_excel_columns(df):
    """Detect enrollment number, student name, and enrollment date columns (NAAC 2.1 layouts)."""
    orig = list(df.columns)
    en_i = nm_i = dt_i = None

    for i, col in enumerate(orig):
        s = _c21_compact_header(col)
        if en_i is None and any(
            k in s
            for k in (
                "studentenrollmentnumber",
                "enrollmentnumber",
                "enrollmentno",
                "enrolmentnumber",
                "enrolmentno",
                "seatno",
                "rollno",
                "prn",
                "registrationno",
                "regno",
                "applicationno",
            )
        ):
            en_i = i
        elif en_i is None and s in ("eno", "en", "roll", "seat"):
            en_i = i

        if nm_i is None and any(
            k in s
            for k in (
                "studentname",
                "nameofstudent",
                "nameofstudents",
                "fullname",
                "candidate",
                "candidatename",
            )
        ):
            nm_i = i
        elif nm_i is None and _c23_normalize_header(col).strip() in ("name", "student"):
            nm_i = i

        if dt_i is None and any(
            k in s
            for k in (
                "dateofenrollment",
                "enrollmentdate",
                "dateofenrolment",
                "enrolmentdate",
                "admissiondate",
                "dateofadmission",
            )
        ):
            dt_i = i

    if en_i is None or nm_i is None:
        if len(orig) >= 2 and en_i is None and nm_i is None:
            en_i, nm_i = 1, 0
        elif len(orig) >= 2 and en_i is None:
            en_i = 0 if nm_i != 0 else 1
        elif len(orig) >= 2 and nm_i is None:
            nm_i = 0 if en_i != 0 else 1

    return en_i, nm_i, dt_i, orig


def _c21_rows_from_df(df):
    en_i, nm_i, dt_i, orig = _c21_pick_excel_columns(df)
    if en_i is None or nm_i is None or en_i == nm_i:
        return [], (
            "Could not detect columns. Use headers such as "
            "Student Enrollment Number, Name of Students, and Date of Enrollment."
        )

    df = df.dropna(how="all")
    if len(df) > _C21_MAX_STUDENTS:
        return [], f"Too many rows (max {_C21_MAX_STUDENTS}). Split the file."

    en_key, nm_key = orig[en_i], orig[nm_i]
    dt_key = orig[dt_i] if dt_i is not None else None

    students = []
    seen: set[str] = set()
    for _, row in df.iterrows():
        en = _c21_clean_cell_str(row.get(en_key))
        nm = _c21_clean_cell_str(row.get(nm_key))
        if not en and not nm:
            continue
        if not en or not nm:
            continue
        if en.casefold() in seen:
            continue
        seen.add(en.casefold())

        dt_val = None
        if dt_key is not None:
            dt_val = _c21_parse_date_cell(row.get(dt_key))

        students.append({
            "enrollment_number": en,
            "student_name": nm,
            "enrollment_date": dt_val.isoformat() if dt_val else None,
        })

    if not students:
        return [], "No data rows found. Each row needs enrollment number and student name."

    return students, None


def _c21_students_from_upload(raw: bytes, filename: str) -> tuple[list, str | None]:
    """Parse student rows from PDF, CSV, or Excel upload for Criterion 2.1."""
    ext = os.path.splitext(filename or "")[1].lower()
    if ext == ".pdf":
        try:
            students, _ = _c21_extract_students_from_pdf_bytes(raw)
        except Exception as e:
            return [], f"Could not read PDF: {str(e)}"
        if not students:
            return [], (
                "No student rows detected in PDF. Use selectable text (not scanned image-only). "
                "Each row needs an enrollment number and student name."
            )
        return students, None

    if ext in {".csv", ".xlsx"}:
        try:
            buf = io.BytesIO(raw)
            if ext == ".csv":
                df = pd.read_csv(buf, encoding="utf-8-sig")
            else:
                df = pd.read_excel(buf, engine="openpyxl")
        except Exception as e:
            return [], f"Could not read spreadsheet: {str(e)}"
        return _c21_rows_from_df(df)

    return [], "Only PDF, CSV, or XLSX files are supported."


def _c21_save_source_file(raw: bytes, filename: str) -> str:
    """Store uploaded source list and return public URL for proof column."""
    target_dir = _canonical_evidence_dir_abs("2_1")
    safe = secure_filename(filename or "student_list")
    if not safe:
        safe = "student_list"
    out_name = f"2_1_source_{uuid.uuid4().hex[:8]}_{safe}"
    out_path = os.path.join(target_dir, out_name)
    with open(out_path, "wb") as fh:
        fh.write(raw)
    return _abs_path_to_public_url(out_path)


@api_bp.route("/records/2_1/bulk-import", methods=["POST"])
def bulk_import_c21():
    """Upload PDF/Excel once with shared enrollment year → save all students to table."""
    enrollment_year = (request.form.get("enrollment_year") or "").strip()
    if not enrollment_year:
        return jsonify({"success": False, "error": "Year of enrollment is required."}), 400

    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded."}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "error": "Empty filename."}), 400

    raw = f.read()
    if not raw:
        return jsonify({"success": False, "error": "Empty file."}), 400
    if len(raw) > _C21_MAX_PDF_BYTES:
        return jsonify({"success": False, "error": "File too large (max 15 MB)."}), 400

    students, err = _c21_students_from_upload(raw, f.filename)
    if err:
        return jsonify({"success": False, "error": err, "students": []}), 400

    enrollment_date = None
    raw_date = (request.form.get("enrollment_date") or "").strip()
    if raw_date:
        try:
            p = raw_date[:10].split("-")
            enrollment_date = date(int(p[0]), int(p[1]), int(p[2]))
        except Exception:
            pass

    try:
        proof_link = _c21_save_source_file(raw, f.filename)
        _save_proof_link_value("2_1", proof_link)
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": f"Could not save proof file: {e}"}), 500

    user_id = session.get("user_id")
    added = skipped = updated = 0
    new_rows = []
    seen: set[str] = set()

    for s in students:
        en = str(s.get("enrollment_number") or "").strip()
        nm = str(s.get("student_name") or "").strip()
        if not en or not nm:
            skipped += 1
            continue
        key = en.casefold()
        if key in seen:
            skipped += 1
            continue
        seen.add(key)

        row_date = s.get("enrollment_date")
        if isinstance(row_date, str) and row_date.strip():
            row_date = _c21_parse_date_cell(row_date)
        elif not isinstance(row_date, date):
            row_date = None
        final_date = row_date or enrollment_date

        dup = C21StudentsDuringYear.query.filter_by(
            enrollment_year=enrollment_year,
            enrollment_number=en,
        ).first()
        if dup:
            if final_date and not dup.enrollment_date:
                dup.enrollment_date = final_date
                dup.updated_by_id = user_id
                updated += 1
                new_rows.append(to_dict(dup))
            else:
                skipped += 1
            continue

        rec = C21StudentsDuringYear(
            enrollment_year=enrollment_year,
            student_name=nm,
            enrollment_number=en,
            enrollment_date=final_date,
            created_by_id=user_id,
            updated_by_id=user_id,
        )
        db.session.add(rec)
        db.session.flush()
        added += 1
        new_rows.append(to_dict(rec))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400

    return jsonify({
        "success": True,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "count": len(students),
        "proof_link": proof_link,
        "data": new_rows,
    })


@api_bp.route('/criteria2/2_1/preview-pdf-text', methods=['POST'])
def preview_criterion_21_pdf_text():
    """Return raw text lines extracted from a PDF so the user can inspect the format."""
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400
    f = request.files['file']
    if not f.filename or os.path.splitext(f.filename)[1].lower() != '.pdf':
        return jsonify({"success": False, "error": "Only PDF files are supported"}), 400
    data = f.read()
    if not data:
        return jsonify({"success": False, "error": "Empty file"}), 400
    if len(data) > _C21_MAX_PDF_BYTES:
        return jsonify({"success": False, "error": "File too large (max 15 MB)"}), 400
    try:
        reader = PdfReader(io.BytesIO(data))
        chunks = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                chunks.append(t)
        full = "\n".join(chunks)
        lines = []
        for block in full.replace("\r", "").split("\n"):
            for seg in block.split("\x0c"):
                seg = seg.strip()
                if seg:
                    lines.append(seg)
        return jsonify({"success": True, "raw_lines": lines[:300], "total_lines": len(lines)})
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read PDF: {str(e)}"}), 400


_C23_MAX_UPLOAD_BYTES = 4 * 1024 * 1024
_C23_MAX_ROWS = 2500


def _c23_normalize_header(c):
    return re.sub(r"\s+", " ", str(c).strip()).lower()


def _c23_pick_enrollment_and_name_cols(df):
    """Pick column indices for enrollment number and student name from messy NAAC-style exports."""
    orig = list(df.columns)
    norms = [_c23_normalize_header(c) for c in orig]

    def compact(i):
        return re.sub(r"[\s._\-/]+", "", norms[i])

    en_i = nm_i = None
    for i in range(len(orig)):
        s = compact(i)
        if en_i is None:
            if any(
                k in s
                for k in (
                    "enrollmentno",
                    "enrolmentno",
                    "enrollmentnumber",
                    "seatno",
                    "rollno",
                    "prn",
                    "registrationno",
                    "regno",
                    "applicationno",
                )
            ):
                en_i = i
            elif s in ("eno", "en", "roll", "seat"):
                en_i = i
        if nm_i is None:
            if s in ("studentname", "nameofstudent", "fullname", "candidate", "candidatename"):
                nm_i = i
            elif norms[i].strip() == "name":
                nm_i = i

    if en_i is None or nm_i is None:
        if len(orig) >= 2 and en_i is None and nm_i is None:
            en_i, nm_i = 0, 1
        elif len(orig) >= 2 and en_i is None:
            en_i = 0 if nm_i != 0 else 1
        elif len(orig) >= 2 and nm_i is None:
            nm_i = 0 if en_i != 0 else 1

    return en_i, nm_i, orig


def _c23_rows_from_df(df):
    en_i, nm_i, orig = _c23_pick_enrollment_and_name_cols(df)
    if en_i is None or nm_i is None or en_i == nm_i:
        return [], "Could not detect columns. Use headers such as Enrollment Number / Seat No and Student Name."

    df = df.dropna(how="all")
    if len(df) > _C23_MAX_ROWS:
        return [], f"Too many rows (max {_C23_MAX_ROWS}). Split the file."

    en_key, nm_key = orig[en_i], orig[nm_i]
    students = []
    seen = set()
    for _, row in df.iterrows():
        en = str(row.get(en_key, "") or "").strip()
        nm = str(row.get(nm_key, "") or "").strip()
        if not en and not nm:
            continue
        if not en or not nm:
            continue
        if en in seen:
            continue
        seen.add(en)
        students.append({"enrollment_number": en, "student_name": nm})

    if not students:
        return [], "No data rows found. Each row needs both enrollment number and student name."

    return students, None


@api_bp.route('/criteria2/2_3/parse-outgoing-list', methods=['POST'])
def parse_criterion_23_outgoing_list():
    """Upload CSV or Excel (.xlsx); returns enrollment_number + student_name rows for Criterion 2.3."""
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded", "students": []}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"success": False, "error": "Empty filename", "students": []}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in {'.csv', '.xlsx'}:
        return jsonify({"success": False, "error": "Only .csv or .xlsx files are supported", "students": []}), 400

    raw = f.read()
    if not raw:
        return jsonify({"success": False, "error": "Empty file", "students": []}), 400
    if len(raw) > _C23_MAX_UPLOAD_BYTES:
        return jsonify({"success": False, "error": "File too large (max 4 MB)", "students": []}), 400

    try:
        buf = io.BytesIO(raw)
        if ext == '.csv':
            df = pd.read_csv(buf, encoding='utf-8-sig')
        else:
            df = pd.read_excel(buf, engine='openpyxl')
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read file: {str(e)}", "students": []}), 400

    students, err = _c23_rows_from_df(df)
    if err:
        return jsonify({"success": False, "error": err, "students": []}), 400

    return jsonify({"success": True, "students": students, "count": len(students)})


# ============================================================
# UPLOAD EVIDENCE (combine images + PDFs into one PDF)
# ============================================================

ALLOWED_PDF_EXT   = {'.pdf'}
ALLOWED_IMAGE_EXT = {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.webp'}

def _criterion_subfolder(criterion: str) -> str:
    """Return a subfolder name like 'criteria1', 'criteria2', etc."""
    parts = criterion.split('_')
    if parts and parts[0].isdigit():
        return f"criteria{parts[0]}"
    return "misc"

def _image_to_pdf_bytes(image_bytes: bytes) -> bytes | None:
    try:
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="PDF")
        return buf.getvalue()
    except Exception:
        return None


def _bytes_look_like_pdf(data: bytes) -> bool:
    return bool(data) and data[:4] == b"%PDF"


def _collect_request_upload_files():
    """Gather every file field from multipart upload (supports multiple PDFs)."""
    collected = []
    seen_ids: set[int] = set()
    for key in ("files", "file", "files[]"):
        for f in request.files.getlist(key):
            fid = id(f)
            if fid not in seen_ids:
                seen_ids.add(fid)
                collected.append(f)
    return collected


def _append_pdf_bytes_to_writer(writer: PdfWriter, data: bytes, label: str) -> tuple[bool, str]:
    try:
        if not data:
            return False, f"{label} (empty)"
        if _bytes_look_like_pdf(data):
            reader = PdfReader(io.BytesIO(data), strict=False)
            writer.append(reader)
            return True, ""
        ext = os.path.splitext(label)[1].lower()
        if ext in ALLOWED_PDF_EXT:
            reader = PdfReader(io.BytesIO(data), strict=False)
            writer.append(reader)
            return True, ""
        if ext in ALLOWED_IMAGE_EXT or not ext:
            pdf_bytes = _image_to_pdf_bytes(data)
            if pdf_bytes:
                reader = PdfReader(io.BytesIO(pdf_bytes), strict=False)
                writer.append(reader)
                return True, ""
        return False, f"{label} (unsupported type)"
    except Exception as e:
        return False, f"{label} ({str(e)[:80]})"


def _append_pdf_file_to_writer(writer: PdfWriter, fp: str, label: str) -> tuple[bool, str]:
    with open(fp, "rb") as fh:
        data = fh.read()
    return _append_pdf_bytes_to_writer(writer, data, label or os.path.basename(fp))


def _canonical_evidence_dir_abs(criterion: str) -> str:
    """Save uploads under <static_folder>/uploads/<criteriaN>/ (stable absolute disk path)."""
    sub = _criterion_subfolder(criterion)
    root = os.path.abspath(os.path.join(current_app.static_folder, "uploads", sub))
    os.makedirs(root, exist_ok=True)
    return root


def _abs_path_to_public_url(abs_file_path: str) -> str:
    """Browser path like /uploads/criteria1/file.pdf."""
    static_root = os.path.abspath(current_app.static_folder)
    fp = os.path.abspath(abs_file_path)
    try:
        rel = os.path.relpath(fp, static_root).replace(os.sep, "/")
    except ValueError:
        rel = os.path.basename(fp).replace(os.sep, "/")
    rel = rel.lstrip("/")
    return "/" + rel


def _proof_link_to_excel_full_url(link: str) -> str:
    """Clickable absolute URL for Excel; uses current request host when exporting from browser."""
    raw = (link or "").strip().replace("\\", "/")
    if not raw:
        return ""
    low = raw.lower()
    if low.startswith("http://") or low.startswith("https://"):
        return raw
    path = raw if raw.startswith("/") else "/" + raw
    try:
        base = request.url_root.rstrip("/")
    except RuntimeError:
        base = os.environ.get("NAAC_PUBLIC_BASE_URL", "http://127.0.0.1:5000").rstrip("/")
    return base + path

@api_bp.route('/proof-link/<criterion>', methods=['GET'])
def get_proof_link(criterion):
    """Return the saved global proof link for any criterion."""
    link = _get_proof_link_value(criterion)
    return jsonify({"success": True, "link": link})


@api_bp.route('/proof-link/<criterion>', methods=['POST'])
def save_proof_link(criterion):
    """Save or update the global proof link for any criterion."""
    data = request.json or {}
    link = (data.get('link') or '').strip()
    try:
        _save_proof_link_value(criterion, link)
        return jsonify({"success": True, "link": link})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@api_bp.route('/proof-link/<criterion>', methods=['DELETE'])
def delete_proof_link(criterion):
    """Clear saved proof link so a new document can be uploaded."""
    try:
        _save_proof_link_value(criterion, '')
        return jsonify({"success": True, "link": None})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@api_bp.route('/records/1_3_3/bulk-add', methods=['POST'])
def bulk_add_c133():
    """Add one 1.3.3 row per student name with shared program fields."""
    data = request.json or {}
    pc = (data.get('programCode') or '').strip()
    pn = (data.get('programName') or '').strip()
    names = data.get('studentNames') or []
    if not pc:
        return jsonify({"success": False, "error": "Program code is required."}), 400
    if not names:
        return jsonify({"success": False, "error": "No student names provided."}), 400

    prog = _resolve_or_create_program(pc, pn)
    if not prog:
        return jsonify({"success": False, "error": "Could not resolve programme."}), 400

    user_id = session.get('user_id')
    added = skipped = failed = 0
    new_rows = []
    seen = set()

    for raw in names:
        nm = ' '.join(str(raw or '').split()).strip()
        if not nm:
            continue
        key = nm.lower()
        if key in seen:
            skipped += 1
            continue
        seen.add(key)

        stu = _resolve_or_create_student_by_name(nm)
        if not stu:
            failed += 1
            continue

        dup = C133Projects.query.filter_by(program_id=prog.id, student_id=stu.id).first()
        if dup:
            skipped += 1
            continue

        rec = C133Projects(
            program_id=prog.id,
            student_id=stu.id,
            created_by_id=user_id,
            updated_by_id=user_id,
        )
        db.session.add(rec)
        db.session.flush()
        added += 1
        new_rows.append(to_dict(rec))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400

    return jsonify({
        "success": True,
        "added": added,
        "skipped": skipped,
        "failed": failed,
        "data": new_rows,
    })


def _resolve_local_proof_file_path(link: str) -> str | None:
    """Locate a proof file on disk from DB value (URL, absolute path, /uploads/…, /static/uploads/…)."""
    if not link or not isinstance(link, str):
        return None

    static_root = os.path.abspath(current_app.static_folder)
    raw = link.strip().replace("\\", "/")

    def _filesystem_candidates(web_path: str) -> list[str]:
        """Turn /uploads/foo into absolute paths under static_folder."""
        wp = web_path.strip().split("?", 1)[0].split("#", 1)[0]
        if not wp.startswith("/"):
            wp = "/" + wp
        pl = wp.lower()
        if pl.startswith("/static/uploads/"):
            wp = "/uploads/" + wp[16:].lstrip("/")
        elif pl.startswith("/static/uploads"):
            wp = "/uploads/" + wp[15:].lstrip("/")
        # Only files we serve from Flask static are under /uploads/…
        if "/uploads/" in wp.lower():
            m = re.search(r"(?i)(/uploads/.+)", wp)
            wp = m.group(1) if m else wp
        elif not pl.startswith("/uploads"):
            return []
        rel = wp.lstrip("/").replace("/", os.sep)
        p1 = os.path.normpath(os.path.join(static_root, rel))
        outs = [p1]
        return outs

    trials: list[str] = []

    if raw.lower().startswith(("http://", "https://")):
        path_only = unquote((urlparse(raw).path or "").strip())
        if path_only:
            trials.extend(_filesystem_candidates(path_only))
    trials.extend(_filesystem_candidates(raw))

    if not raw.lower().startswith(("http://", "https://", "//")):
        fs_guess = os.path.normpath(raw.replace("/", os.sep))
        drive_letter = len(raw) > 2 and raw[1] == ":"
        if os.path.isabs(fs_guess) or drive_letter:
            trials.append(os.path.abspath(fs_guess))

    seen: set[str] = set()
    for cand in trials:
        cand = os.path.normpath(os.path.abspath(cand))
        if cand in seen:
            continue
        seen.add(cand)
        try:
            common = os.path.commonpath([static_root, cand])
        except ValueError:
            continue
        if common != static_root:
            continue
        if os.path.isfile(cand):
            return cand

    # Absolute Windows/Unix path stored in DB (file may live outside static/uploads)
    if not raw.lower().startswith(("http://", "https://", "//")):
        abs_guess = os.path.normpath(raw.replace("/", os.sep))
        if os.path.isabs(abs_guess) and os.path.isfile(abs_guess):
            return os.path.abspath(abs_guess)

    # Relative path like uploads/criteria1/file.pdf
    if not raw.lower().startswith(("http://", "https://", "//")) and "/" in raw:
        rel_guess = os.path.normpath(os.path.join(static_root, raw.lstrip("/\\")))
        if os.path.isfile(rel_guess):
            return rel_guess
        upload_folder = current_app.config.get("UPLOAD_FOLDER")
        if upload_folder:
            rel_upload = os.path.normpath(os.path.join(upload_folder, os.path.basename(raw)))
            if os.path.isfile(rel_upload):
                return rel_upload

    # Legacy rows may store bare filename only — search under static/uploads
    base = os.path.basename(raw.replace("\\", "/"))
    if (
        base
        and "/" not in raw.strip().replace("\\", "/")
        and not raw.lower().startswith(("http://", "https://"))
    ):
        uploads_root = os.path.join(static_root, "uploads")
        if os.path.isdir(uploads_root):
            for dirpath, _, filenames in os.walk(uploads_root):
                for fn in filenames:
                    if fn.casefold() == base.casefold():
                        cand = os.path.normpath(os.path.join(dirpath, fn))
                        if os.path.isfile(cand):
                            return cand

    return None


_RECORD_PROOF_ATTRS = (
    "proof_document",
    "proof_links",
    "document_link",
    "supporting_document",
    "doc_link",
)


def _record_proof_path_value(rec):
    """First non-empty per-row proof path/link on a criterion record."""
    for attr in _RECORD_PROOF_ATTRS:
        if hasattr(rec, attr):
            val = getattr(rec, attr, None)
            if val and str(val).strip():
                return str(val).strip()
    return ""


@api_bp.route('/combine-record-proofs/<criterion>', methods=['POST'])
def combine_record_proofs(criterion):
    """Merge per-row proof files already saved on records → one PDF for Excel export."""
    model = CRITERIA_MODELS.get(criterion)
    if not model:
        return jsonify({"success": False, "error": f"Unknown criterion: {criterion}"}), 404

    session_dept = ""
    uid = session.get("user_id")
    if uid:
        u = User.query.get(uid)
        if u:
            session_dept = (u.department or "").strip()

    records = model.query.order_by(model.id).all()
    writer = PdfWriter()
    merged = 0
    skipped = []

    for rec in records:
        if criterion == "1_1_3" and session_dept and getattr(rec, "teacher_id", None):
            tc = Teacher.query.get(rec.teacher_id)
            if tc:
                td = (tc.department or "").strip().casefold()
                if td and td != session_dept.casefold():
                    continue

        pv = _record_proof_path_value(rec)
        if not pv:
            continue

        fp = _resolve_local_proof_file_path(pv)
        if not fp:
            skipped.append(
                f"Row id {rec.id}: proof missing on server — re-save row with PDF/image attached"
            )
            continue

        label = f"Row id {rec.id}"
        ok, err = _append_pdf_file_to_writer(writer, fp, os.path.basename(fp))
        if ok:
            merged += 1
        else:
            skipped.append(f"{label}: {err}")

    if merged == 0:
        return jsonify({
            "success": False,
            "error": "No row proofs found on disk to merge. Attach PDF/image on each record, Save, then combine.",
            "skipped": skipped,
            "rows_checked": len(records),
        }), 400

    target_dir = _canonical_evidence_dir_abs(criterion)
    out_name = f"{criterion}_combined_{uuid.uuid4().hex[:8]}.pdf"
    out_path = os.path.join(target_dir, out_name)
    with open(out_path, "wb") as out_f:
        writer.write(out_f)

    public_link = _abs_path_to_public_url(out_path)

    try:
        _save_proof_link_value(criterion, public_link)
    except Exception as ex:
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": f"Merged file written but could not save proof link: {ex}",
        }), 500

    return jsonify({
        "success": True,
        "link": public_link,
        "merged_from_records": merged,
        "skipped": skipped,
    })


@api_bp.route('/upload-evidence', methods=['POST'])
def upload_evidence():
    """Combine multiple uploaded files (PDFs + images) into a single PDF.
    Returns: { success, link, combined_count, skipped }
    """
    criterion = request.form.get('criterion', '').strip()
    if not criterion:
        return jsonify({"success": False, "error": "criterion is required"}), 400

    files = _collect_request_upload_files()
    if not files:
        return jsonify({"success": False, "error": "No files uploaded"}), 400

    target_dir = _canonical_evidence_dir_abs(criterion)

    writer = PdfWriter()
    combined_count = 0
    merged_names = []
    skipped = []

    for f in files:
        original_name = secure_filename(f.filename or f"file_{combined_count + 1}.pdf")
        data = f.read()
        ok, err = _append_pdf_bytes_to_writer(writer, data, original_name)
        if ok:
            combined_count += 1
            merged_names.append(original_name)
        else:
            skipped.append(err)

    if combined_count == 0:
        return jsonify({
            "success": False,
            "error": "No valid PDFs/images could be combined.",
            "skipped": skipped,
        }), 400

    out_name = f"{criterion}_combined_{uuid.uuid4().hex[:8]}.pdf"
    out_path = os.path.join(target_dir, out_name)
    with open(out_path, 'wb') as out_file:
        writer.write(out_file)

    public_link = _abs_path_to_public_url(out_path)
    page_count = len(PdfReader(out_path, strict=False).pages)

    return jsonify({
        "success": True,
        "link": public_link,
        "combined_count": combined_count,
        "merged_files": merged_names,
        "page_count": page_count,
        "skipped": skipped,
    })

register_c346_routes(api_bp, to_dict)
